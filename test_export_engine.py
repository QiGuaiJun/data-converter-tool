from __future__ import annotations

import json
from pathlib import Path
from io import BytesIO

from openpyxl import load_workbook

import server


def sqlite_fixture() -> None:
    with server.connect_db() as conn:
        conn.execute("drop table if exists export_people")
        conn.execute("create table export_people (name text, amount integer, city text)")
        conn.executemany(
            "insert into export_people values (?, ?, ?)",
            [("Alice", 10, "北京"), ("Bob", 20, "上海"), ("Carol", 30, "北京")],
        )


def assert_xlsx(path: str, expected_rows: int) -> None:
    workbook = load_workbook(path)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    workbook.close()
    assert len(rows) == expected_rows


def test_sqlite_exports() -> None:
    sqlite_fixture()
    result = server.run_export_job(
        {
            "targetDbType": "sqlite",
            "items": [{"type": "table", "table": "export_people", "name": "export_people"}],
            "extension": "xlsx",
            "outputName": "qa_export_people",
            "sheetName": "People",
            "exportMode": "workbook",
            "headerMode": "field",
            "exportFields": "name,amount",
            "whereClause": "amount >= 20",
            "rowHeight": "22",
            "columnWidth": "18",
            "fontName": "Arial",
            "fontSize": "11",
            "addBorder": "true",
            "lockHeader": "true",
        }
    )
    assert result["rows"] == 2
    assert_xlsx(result["files"][0], 3)

    csv_result = server.run_export_job(
        {
            "targetDbType": "sqlite",
            "items": [{"type": "query", "name": "query_amount", "sql": "select name, amount from export_people order by amount"}],
            "extension": "csv",
            "outputName": "qa_export_people_csv",
            "headerMode": "field",
            "encoding": "utf-8",
            "delimiter": ",",
        }
    )
    assert Path(csv_result["files"][0]).read_text(encoding="utf-8").startswith("name,amount")

    json_result = server.run_export_job(
        {
            "targetDbType": "sqlite",
            "items": [{"type": "table", "table": "export_people", "name": "export_people"}],
            "extension": "json",
            "outputName": "qa_export_people_json",
            "splitField": "city",
        }
    )
    assert len(json_result["files"]) == 2
    assert json.loads(Path(json_result["files"][0]).read_text(encoding="utf-8"))

    xml_result = server.run_export_job(
        {
            "targetDbType": "sqlite",
            "items": [{"type": "query", "name": "query_xml", "sql": "select name from export_people where city = '北京'"}],
            "extension": "xml",
            "outputName": "qa_export_people_xml",
        }
    )
    assert Path(xml_result["files"][0]).read_text(encoding="utf-8").startswith("<?xml")


def test_mysql_export_if_available() -> None:
    with server.connect_db() as conn:
        row = conn.execute("select id from _db_connections where db_type = 'mysql' order by updated_at desc limit 1").fetchone()
    if not row:
        return

    fields = {
        "connectionId": row["id"],
        "importMode": "rebuild",
        "tableName": "codex_export_people",
        "mapping": json.dumps(
            [
                {"sourceIndex": 0, "target": "Name", "enabled": True, "defaultValue": "", "matchKey": True},
                {"sourceIndex": 1, "target": "Amount", "enabled": True, "defaultValue": "0", "matchKey": False},
            ]
        ),
        "tableCase": "lower",
        "fieldCase": "lower",
        "commitMode": "once",
    }
    upload = Path("uploads/codex_export_people.csv")
    upload.write_text("Name,Amount\nA,1\nB,2\n", encoding="utf-8-sig")
    server.import_uploaded_file(server.UploadedFile("codex_export_people.csv", upload), fields)

    result = server.run_export_job(
        {
            "connectionId": row["id"],
            "targetDbType": "mysql",
            "items": [{"type": "table", "table": "codex_export_people", "name": "codex_export_people"}],
            "extension": "xlsx",
            "outputName": "qa_mysql_export_people",
            "sheetName": "MySQL",
        }
    )
    assert result["rows"] == 2
    assert_xlsx(result["files"][0], 3)

    conn = server.connect_target_db({"connectionId": row["id"]})
    try:
        with conn.cursor() as cursor:
            cursor.execute("drop table if exists codex_export_people")
        conn.commit()
    finally:
        conn.close()


def test_large_streaming_export() -> None:
    with server.connect_db() as conn:
        conn.execute("drop table if exists export_large_people")
        conn.execute("create table export_large_people (id integer, name text, amount integer)")
        conn.executemany(
            "insert into export_large_people values (?, ?, ?)",
            ((index, f"name_{index}", index % 100) for index in range(20000)),
        )

    xlsx_result = server.run_export_job(
        {
            "targetDbType": "sqlite",
            "items": [{"type": "table", "table": "export_large_people", "name": "export_large_people"}],
            "extension": "xlsx",
            "outputName": "qa_large_streaming_xlsx",
            "sheetName": "Large",
            "exportMode": "workbook",
            "headerMode": "field",
        }
    )
    assert xlsx_result["rows"] == 20000
    assert_xlsx(xlsx_result["files"][0], 20001)

    csv_result = server.run_export_job(
        {
            "targetDbType": "sqlite",
            "items": [{"type": "table", "table": "export_large_people", "name": "export_large_people"}],
            "extension": "csv",
            "outputName": "qa_large_streaming_csv",
            "headerMode": "field",
        }
    )
    assert csv_result["rows"] == 20000
    assert Path(csv_result["files"][0]).read_text(encoding="utf-8").splitlines()[0] == "id,name,amount"


class FakeDownloadHandler:
    def __init__(self) -> None:
        self.headers: list[tuple[str, str]] = []
        self.status = None
        self.wfile = BytesIO()

    def send_response(self, status: int) -> None:
        self.status = status

    def send_header(self, key: str, value: str) -> None:
        value.encode("latin-1")
        self.headers.append((key, value))

    def end_headers(self) -> None:
        pass


def test_chinese_filename_download_header() -> None:
    path = server.EXPORTS / "\u4e2d\u6587\u5bfc\u51fa\u6587\u4ef6.xlsx"
    path.write_bytes(b"demo")
    handler = FakeDownloadHandler()
    server.ImportPrototypeHandler.handle_export_download(handler, "name=%E4%B8%AD%E6%96%87%E5%AF%BC%E5%87%BA%E6%96%87%E4%BB%B6.xlsx")
    assert handler.status == 200
    disposition = dict(handler.headers)["Content-Disposition"]
    assert disposition == "attachment; filename=\"export.xlsx\"; filename*=UTF-8''%E4%B8%AD%E6%96%87%E5%AF%BC%E5%87%BA%E6%96%87%E4%BB%B6.xlsx"



if __name__ == "__main__":
    test_sqlite_exports()
    test_mysql_export_if_available()
    test_large_streaming_export()
    test_chinese_filename_download_header()
    print("export engine checks passed")
