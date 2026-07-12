from __future__ import annotations

import csv
import base64
import datetime as dt
import hmac
import json
import mimetypes
import os
import re
import sqlite3
import tempfile
import threading
import time
import uuid
import xml.etree.ElementTree as ET
import zipfile
from collections.abc import Iterable, Iterator
from contextlib import nullcontext
from dataclasses import dataclass
from concurrent.futures import ThreadPoolExecutor
from itertools import chain
from io import BytesIO
from http import HTTPStatus
from http.server import SimpleHTTPRequestHandler, ThreadingHTTPServer
from pathlib import Path
from urllib.parse import parse_qs, quote, urlparse
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import msoffcrypto
import pymysql
import xlrd
from dbfread import DBF
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Border, Font, Protection, Side
from openpyxl.utils import get_column_letter
from pypinyin import Style, lazy_pinyin
from xml.sax.saxutils import escape as xml_escape


ROOT = Path(__file__).resolve().parent
PUBLIC = ROOT / "public"


def env_path(name: str, fallback: Path) -> Path:
    raw_value = os.environ.get(name, "").strip()
    volume_mount = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", "").strip()
    if raw_value:
        path = Path(raw_value)
        if volume_mount and not path.is_absolute():
            return (Path(volume_mount) / path).resolve()
        return path.resolve()
    if volume_mount:
        return (Path(volume_mount) / fallback.name).resolve()
    return fallback.resolve()


DATA = env_path("DATA_DIR", ROOT / "data")
UPLOADS = env_path("UPLOADS_DIR", ROOT / "uploads")
EXPORTS = env_path("EXPORTS_DIR", ROOT / "exports")
DB_PATH = DATA / "imports.db"
PASSWORD_PREFIX = "b64:"

MAX_PREVIEW_ROWS = 20
EXPORT_FETCH_SIZE = 5000
SUPPORTED_EXTENSIONS = {".csv", ".txt", ".xlsx", ".xlsm", ".xls", ".json", ".xml", ".dbf"}


@dataclass
class UploadedFile:
    filename: str
    path: Path


@dataclass
class TabularData:
    columns: list[str]
    rows: list[list[str]]
    sheets: list[str]
    selected_sheet: str


def ensure_dirs() -> None:
    DATA.mkdir(parents=True, exist_ok=True)
    UPLOADS.mkdir(parents=True, exist_ok=True)
    EXPORTS.mkdir(parents=True, exist_ok=True)


def connect_db() -> sqlite3.Connection:
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH, timeout=30)
    conn.row_factory = sqlite3.Row
    conn.execute("pragma busy_timeout = 30000")
    conn.execute(
        """
        create table if not exists _import_logs (
            id text primary key,
            created_at text not null,
            file_name text not null,
            table_name text not null,
            mode text not null,
            rows_read integer not null default 0,
            rows_written integer not null,
            rows_updated integer not null default 0,
            rows_skipped integer not null default 0,
            status text not null,
            message text not null
        )
        """
    )
    conn.execute(
        """
        create table if not exists _import_checkpoints (
            key text primary key,
            updated_at text not null,
            rows_done integer not null
        )
        """
    )
    conn.execute(
        """
        create table if not exists _db_connections (
            id text primary key,
            name text not null,
            db_type text not null,
            host text not null,
            port integer not null,
            user_name text not null,
            password text not null default '',
            db_name text not null default '',
            charset text not null default 'utf8mb4',
            ssl_enabled integer not null default 0,
            ssl_ca text not null default '',
            ssl_cert text not null default '',
            ssl_key text not null default '',
            created_at text not null,
            updated_at text not null
        )
        """
    )
    conn.execute(
        """
        create table if not exists _jobs (
            id text primary key,
            name text not null,
            enabled integer not null default 1,
            steps_json text not null default '[]',
            created_at text not null,
            updated_at text not null
        )
        """
    )
    conn.execute(
        """
        create table if not exists _schedules (
            id text primary key,
            name text not null,
            job_id text not null,
            enabled integer not null default 0,
            rule_json text not null default '{}',
            start_at text not null default '',
            end_at text not null default '',
            next_run_at text not null default '',
            last_run_at text not null default '',
            last_status text not null default '',
            log_retention_days integer not null default 3,
            email_on_fail integer not null default 0,
            running integer not null default 0,
            created_at text not null,
            updated_at text not null
        )
        """
    )
    conn.execute(
        """
        create table if not exists _job_runs (
            id text primary key,
            job_id text not null,
            schedule_id text not null default '',
            job_name text not null,
            started_at text not null,
            ended_at text not null default '',
            elapsed_ms integer not null default 0,
            status text not null,
            message text not null default ''
        )
        """
    )
    conn.execute(
        """
        create table if not exists _job_run_steps (
            id text primary key,
            run_id text not null,
            step_index integer not null,
            step_name text not null,
            step_type text not null,
            started_at text not null,
            ended_at text not null default '',
            elapsed_ms integer not null default 0,
            status text not null,
            message text not null default ''
        )
        """
    )
    for column, ddl in {
        "rows_read": "integer not null default 0",
        "rows_updated": "integer not null default 0",
        "rows_skipped": "integer not null default 0",
    }.items():
        existing = [row["name"] for row in conn.execute("pragma table_info(_import_logs)").fetchall()]
        if column not in existing:
            conn.execute(f"alter table _import_logs add column {column} {ddl}")
    return conn


def encode_secret(value: str) -> str:
    if not value:
        return ""
    return PASSWORD_PREFIX + base64.b64encode(value.encode("utf-8")).decode("ascii")


def decode_secret(value: str) -> str:
    if not value:
        return ""
    if not value.startswith(PASSWORD_PREFIX):
        return value
    try:
        return base64.b64decode(value[len(PASSWORD_PREFIX) :]).decode("utf-8")
    except Exception:
        return ""


def normalize_connection_payload(payload: dict[str, object]) -> dict[str, object]:
    db_type = str(payload.get("dbType") or payload.get("targetDbType") or "mysql").strip().lower()
    if db_type != "mysql":
        raise ValueError("当前连接模块先支持 MySQL，其他数据库会在后续模块开放。")
    host = str(payload.get("host") or payload.get("dbHost") or "").strip()
    user = str(payload.get("user") or payload.get("dbUser") or "").strip()
    if not host:
        raise ValueError("请填写主机。")
    if not user:
        raise ValueError("请填写用户名。")
    name = str(payload.get("name") or payload.get("connectionName") or "").strip()
    database = str(payload.get("database") or payload.get("dbName") or "").strip()
    if not name:
        name = f"MySQL - {host}{('/' + database) if database else ''}"
    return {
        "id": str(payload.get("id") or uuid.uuid4().hex),
        "name": name,
        "db_type": db_type,
        "host": host,
        "port": int(payload.get("port") or payload.get("dbPort") or 3306),
        "user_name": user,
        "password": str(payload.get("password") or payload.get("dbPassword") or ""),
        "db_name": database,
        "charset": str(payload.get("charset") or payload.get("dbCharset") or "utf8mb4").strip() or "utf8mb4",
        "ssl_enabled": 1 if payload.get("sslEnabled") in (True, "true", "1", 1, "on") else 0,
        "ssl_ca": str(payload.get("sslCa") or ""),
        "ssl_cert": str(payload.get("sslCert") or ""),
        "ssl_key": str(payload.get("sslKey") or ""),
    }


def connection_public(row: sqlite3.Row, include_password: bool = False) -> dict[str, object]:
    item = {
        "id": row["id"],
        "name": row["name"],
        "dbType": row["db_type"],
        "host": row["host"],
        "port": row["port"],
        "user": row["user_name"],
        "database": row["db_name"],
        "charset": row["charset"],
        "sslEnabled": bool(row["ssl_enabled"]),
        "sslCa": row["ssl_ca"],
        "sslCert": row["ssl_cert"],
        "sslKey": row["ssl_key"],
        "hasPassword": bool(row["password"]),
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }
    if include_password:
        item["password"] = decode_secret(row["password"])
    return item


def load_saved_connection(connection_id: str) -> dict[str, object]:
    if not connection_id:
        return {}
    with connect_db() as conn:
        row = conn.execute("select * from _db_connections where id = ?", (connection_id,)).fetchone()
    if not row:
        raise ValueError("选择的数据库连接不存在，请重新选择。")
    item = connection_public(row, include_password=True)
    return {
        "targetDbType": item["dbType"],
        "dbHost": item["host"],
        "dbPort": str(item["port"]),
        "dbUser": item["user"],
        "dbPassword": item.get("password", ""),
        "dbName": item["database"],
        "dbCharset": item["charset"],
        "sslEnabled": "true" if item["sslEnabled"] else "false",
        "sslCa": item["sslCa"],
        "sslCert": item["sslCert"],
        "sslKey": item["sslKey"],
    }


def resolve_connection_fields(fields: dict[str, str]) -> dict[str, str]:
    connection_id = fields.get("connectionId", "").strip()
    if not connection_id:
        return fields
    saved = load_saved_connection(connection_id)
    merged = dict(fields)
    merged.update({key: str(value) for key, value in saved.items()})
    return merged


def target_db_type(fields: dict[str, str]) -> str:
    fields = resolve_connection_fields(fields)
    return fields.get("targetDbType", "sqlite").strip().lower() or "sqlite"


def connect_target_db(fields: dict[str, str]):
    fields = resolve_connection_fields(fields)
    if target_db_type(fields) == "mysql":
        ssl_config = None
        if parse_bool(fields, "sslEnabled", False):
            ssl_config = {}
            if fields.get("sslCa"):
                ssl_config["ca"] = fields["sslCa"]
            if fields.get("sslCert"):
                ssl_config["cert"] = fields["sslCert"]
            if fields.get("sslKey"):
                ssl_config["key"] = fields["sslKey"]
        return pymysql.connect(
            host=fields.get("dbHost", "127.0.0.1"),
            port=parse_int(fields, "dbPort", 3306),
            user=fields.get("dbUser", ""),
            password=fields.get("dbPassword", ""),
            database=fields.get("dbName", ""),
            charset=fields.get("dbCharset", "utf8mb4") or "utf8mb4",
            autocommit=fields.get("commitMode") == "auto",
            local_infile=fields.get("writeMode") == "load",
            ssl=ssl_config,
        )
    return connect_db()


def db_placeholder(fields: dict[str, str]) -> str:
    return "%s" if target_db_type(fields) == "mysql" else "?"


def db_quote(name: str, fields: dict[str, str]) -> str:
    if target_db_type(fields) == "mysql":
        return "`" + name.replace("`", "``") + "`"
    return '"' + name.replace('"', '""') + '"'


def db_now_sql(fields: dict[str, str]) -> str:
    return "now()" if target_db_type(fields) == "mysql" else "datetime('now', 'localtime')"


def fetch_all_dicts(cursor) -> list[dict[str, object]]:
    columns = [desc[0] for desc in cursor.description or []]
    return [dict(zip(columns, row)) for row in cursor.fetchall()]


def mysql_connection_args(config: dict[str, object], include_database: bool = True) -> dict[str, object]:
    args: dict[str, object] = {
        "host": str(config.get("host") or config.get("dbHost") or "127.0.0.1"),
        "port": int(config.get("port") or config.get("dbPort") or 3306),
        "user": str(config.get("user") or config.get("user_name") or config.get("dbUser") or ""),
        "password": str(config.get("password") or config.get("dbPassword") or ""),
        "charset": str(config.get("charset") or config.get("dbCharset") or "utf8mb4") or "utf8mb4",
        "connect_timeout": 6,
        "read_timeout": 10,
        "write_timeout": 10,
        "local_infile": config.get("writeMode") == "load",
        "autocommit": True,
    }
    database = str(config.get("database") or config.get("dbName") or config.get("db_name") or "").strip()
    if include_database and database:
        args["database"] = database
    ssl_enabled = config.get("sslEnabled") in (True, "true", "1", 1, "on")
    if ssl_enabled:
        ssl_config = {}
        for source, target in (("sslCa", "ca"), ("sslCert", "cert"), ("sslKey", "key")):
            value = str(config.get(source) or "").strip()
            if value:
                ssl_config[target] = value
        args["ssl"] = ssl_config
    return args


def test_mysql_connection(config: dict[str, object]) -> dict[str, object]:
    normalized = normalize_connection_payload(config)
    with pymysql.connect(**mysql_connection_args(normalized, include_database=False)) as conn:
        with conn.cursor() as cursor:
            cursor.execute("select version()")
            version = cursor.fetchone()[0]
            cursor.execute("show databases")
            databases = [row[0] for row in cursor.fetchall()]
    selected_db = str(normalized.get("db_name") or "")
    if selected_db:
        with pymysql.connect(**mysql_connection_args(normalized, include_database=True)) as conn:
            with conn.cursor() as cursor:
                cursor.execute("select database()")
                cursor.fetchone()
        if selected_db not in databases:
            databases.insert(0, selected_db)
    return {"version": version, "databases": databases}


def checkpoint_key(uploaded: UploadedFile, table_name: str, fields: dict[str, str]) -> str:
    stat = uploaded.path.stat()
    return f"{uploaded.filename}|{stat.st_size}|{int(stat.st_mtime)}|{target_db_type(fields)}|{table_name}|{fields.get('importMode', 'append')}"


def get_checkpoint(key: str) -> int:
    with connect_db() as conn:
        row = conn.execute("select rows_done from _import_checkpoints where key = ?", (key,)).fetchone()
        return int(row["rows_done"]) if row else 0


def set_checkpoint(key: str, rows_done: int) -> None:
    with connect_db() as conn:
        conn.execute(
            """
            insert into _import_checkpoints (key, updated_at, rows_done)
            values (?, datetime('now', 'localtime'), ?)
            on conflict(key) do update set updated_at = excluded.updated_at, rows_done = excluded.rows_done
            """,
            (key, rows_done),
        )


def clear_checkpoint(key: str) -> None:
    with connect_db() as conn:
        conn.execute("delete from _import_checkpoints where key = ?", (key,))


def json_response(handler: SimpleHTTPRequestHandler, payload: object, status: int = 200) -> None:
    body = json.dumps(payload, ensure_ascii=False, indent=2).encode("utf-8")
    handler.send_response(status)
    handler.send_header("Content-Type", "application/json; charset=utf-8")
    handler.send_header("Content-Length", str(len(body)))
    handler.end_headers()
    handler.wfile.write(body)


def error_response(handler: SimpleHTTPRequestHandler, message: str, status: int = 400) -> None:
    json_response(handler, {"ok": False, "error": message}, status)


def parse_bool(fields: dict[str, str], name: str, default: bool = False) -> bool:
    value = fields.get(name)
    if value is None:
        return default
    return value.lower() in {"1", "true", "yes", "on"}


def parse_int(fields: dict[str, str], name: str, default: int = 0) -> int:
    value = fields.get(name, "").strip()
    if not value:
        return default
    try:
        return int(value)
    except ValueError as exc:
        raise ValueError(f"{name} 必须是整数。") from exc


def split_values(value: str) -> list[str]:
    return [item.strip() for item in re.split(r"[,，\n]", value or "") if item.strip()]


def decode_escaped(value: str) -> str:
    if not value:
        return ""
    return value.encode("utf-8").decode("unicode_escape")


def sanitize_identifier(value: str, fallback: str) -> str:
    text = (value or "").strip()
    text = re.sub(r"[^\w\u4e00-\u9fff]+", "_", text, flags=re.UNICODE)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = fallback
    if text[0].isdigit():
        text = f"t_{text}"
    return text[:60]


def to_pinyin_initials(value: str) -> str:
    return "".join(lazy_pinyin(value, style=Style.FIRST_LETTER))


def unique_names(names: list[str]) -> list[str]:
    counts: dict[str, int] = {}
    result: list[str] = []
    for index, name in enumerate(names, start=1):
        base = sanitize_identifier(name, f"column_{index}")
        count = counts.get(base, 0)
        counts[base] = count + 1
        result.append(base if count == 0 else f"{base}_{count + 1}")
    return result


def cell_to_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, dt.datetime):
        return value.isoformat(sep=" ", timespec="seconds")
    if isinstance(value, dt.date):
        return value.isoformat()
    return str(value)


def trim_trailing_blanks(row: list[str]) -> list[str]:
    end = len(row)
    while end > 0 and row[end - 1].strip() == "":
        end -= 1
    return row[:end]


def read_csv_rows(path: Path, encoding_option: str, delimiter: str, line_delimiter: str = "") -> list[list[str]]:
    encodings = ["utf-8-sig", "utf-8", "gbk", "gb18030"] if encoding_option == "auto" else [encoding_option]
    last_error: Exception | None = None
    for encoding in encodings:
        try:
            with path.open("r", encoding=encoding, newline="") as file:
                content = file.read()
                sample = content[:4096]
                source_lines = content.split(line_delimiter) if line_delimiter else content.splitlines()
                if delimiter:
                    dialect = csv.excel
                    dialect.delimiter = delimiter
                else:
                    try:
                        dialect = csv.Sniffer().sniff(sample)
                    except csv.Error:
                        dialect = csv.excel
                return [[cell_to_text(cell) for cell in row] for row in csv.reader(source_lines, dialect)]
        except UnicodeDecodeError as exc:
            last_error = exc
    raise ValueError(f"无法识别 CSV/TXT 编码：{last_error}")


def read_excel_rows(path: Path, fields: dict[str, str]) -> tuple[list[list[str]], list[str], str]:
    source = path
    decrypted: BytesIO | None = None
    password = fields.get("excelPassword", "").strip()
    if password:
        decrypted = BytesIO()
        with path.open("rb") as file:
            office_file = msoffcrypto.OfficeFile(file)
            office_file.load_key(password=password)
            office_file.decrypt(decrypted)
        decrypted.seek(0)
        source = decrypted  # type: ignore[assignment]
    workbook = load_workbook(source, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    if not sheet_names:
        workbook.close()
        raise ValueError("Excel 文件没有工作表。")

    mode = fields.get("sheetFilterMode", "name")
    requested = fields.get("sheetName", "").strip()
    selected = sheet_names[0]
    if requested:
        if mode == "index":
            index = int(requested) - 1
            if index < 0 or index >= len(sheet_names):
                workbook.close()
                raise ValueError("指定的 Sheet 序号不存在。")
            selected = sheet_names[index]
        elif requested in sheet_names:
            selected = requested
        else:
            workbook.close()
            raise ValueError("指定的 Sheet 名称不存在。")

    sheet = workbook[selected]
    rows = [[cell_to_text(cell) for cell in row] for row in sheet.iter_rows(values_only=True)]
    workbook.close()
    return rows, sheet_names, selected


def read_xls_rows(path: Path, fields: dict[str, str]) -> tuple[list[list[str]], list[str], str]:
    workbook = xlrd.open_workbook(path)
    sheet_names = workbook.sheet_names()
    if not sheet_names:
        raise ValueError("Excel 文件没有工作表。")
    mode = fields.get("sheetFilterMode", "name")
    requested = fields.get("sheetName", "").strip()
    selected = sheet_names[0]
    if requested:
        if mode == "index":
            index = int(requested) - 1
            if index < 0 or index >= len(sheet_names):
                raise ValueError("指定的 Sheet 序号不存在。")
            selected = sheet_names[index]
        elif requested in sheet_names:
            selected = requested
        else:
            raise ValueError("指定的 Sheet 名称不存在。")
    sheet = workbook.sheet_by_name(selected)
    rows = [[cell_to_text(sheet.cell_value(row, col)) for col in range(sheet.ncols)] for row in range(sheet.nrows)]
    return rows, sheet_names, selected


def read_dbf_rows(path: Path, fields: dict[str, str]) -> list[list[str]]:
    encoding = fields.get("encoding", "auto")
    kwargs = {} if encoding == "auto" else {"encoding": encoding}
    table = DBF(str(path), load=True, char_decode_errors="ignore", **kwargs)
    columns = list(table.field_names)
    rows = [columns]
    for record in table:
        rows.append([cell_to_text(record.get(column, "")) for column in columns])
    return rows


def flatten_object(data: dict[str, object], prefix: str = "") -> dict[str, object]:
    result: dict[str, object] = {}
    for key, value in data.items():
        name = f"{prefix}_{key}" if prefix else str(key)
        if isinstance(value, dict):
            for child_key, child_value in value.items():
                result[f"{name}_{child_key}"] = child_value
        elif isinstance(value, list):
            result[name] = json.dumps(value, ensure_ascii=False)
        else:
            result[name] = value
    return result


def read_json_rows(path: Path, row_tag: str = "") -> list[list[str]]:
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(data, dict):
        data = data.get(row_tag) if row_tag and row_tag in data else next((v for v in data.values() if isinstance(v, list)), [data])
    if not isinstance(data, list):
        raise ValueError("JSON 需要是对象数组，或包含对象数组的根对象。")
    flattened = [flatten_object(item) for item in data if isinstance(item, dict)]
    keys: list[str] = []
    for item in flattened:
        if isinstance(item, dict):
            for key in item.keys():
                if key not in keys:
                    keys.append(str(key))
    if not keys:
        raise ValueError("JSON 中没有可导入的对象行。")
    rows = [keys]
    for item in flattened:
        if isinstance(item, dict):
            rows.append([cell_to_text(item.get(key, "")) for key in keys])
    return rows


def read_xml_rows(path: Path, row_tag: str) -> list[list[str]]:
    root = ET.parse(path).getroot()
    elements = root.findall(f".//{row_tag}") if row_tag else list(root)
    if not elements:
        raise ValueError("XML 中没有可导入的行节点。")
    flattened_rows: list[dict[str, object]] = []
    for element in elements:
        row: dict[str, object] = {}
        for child in list(element):
            if list(child):
                for grandchild in list(child):
                    row[f"{child.tag}_{grandchild.tag}"] = grandchild.text or ""
            else:
                row[child.tag] = child.text or ""
        flattened_rows.append(row)
    keys: list[str] = []
    for row in flattened_rows:
        for key in row.keys():
            if key not in keys:
                keys.append(key)
    if not keys:
        raise ValueError("XML 行节点中没有字段。")
    rows = [keys]
    for row in flattened_rows:
        rows.append([cell_to_text(row.get(key, "")) for key in keys])
    return rows


def make_tabular(raw_rows: list[list[str]], fields: dict[str, str], sheets: list[str] | None = None, selected_sheet: str = "") -> TabularData:
    delete_empty_rows = parse_bool(fields, "deleteEmptyRows", True)
    rows = [trim_trailing_blanks(row) for row in raw_rows]
    if delete_empty_rows:
        rows = [row for row in rows if any(cell.strip() for cell in row)]
    if not rows:
        raise ValueError("文件中没有可导入的数据。")

    header_row = max(parse_int(fields, "headerRow", 1), 1)
    header_index = header_row - 1
    if header_index >= len(rows):
        raise ValueError("表头所在行号超出文件行数。")

    has_header = parse_bool(fields, "hasHeader", True)
    width = max(len(row) for row in rows)
    rows = [row + [""] * (width - len(row)) for row in rows]
    if has_header:
        columns = unique_names(rows[header_index])
        default_data_start = header_row + 1
    else:
        columns = [f"column_{index}" for index in range(1, width + 1)]
        default_data_start = header_row

    data_start = parse_int(fields, "dataStartRow", default_data_start) or default_data_start
    data_rows = rows[max(data_start - 1, 0) :]
    skip_tail = max(parse_int(fields, "skipTailRows", 0), 0)
    if skip_tail:
        data_rows = data_rows[:-skip_tail] if skip_tail < len(data_rows) else []
    import_count = parse_int(fields, "importRowCount", 0)
    if import_count > 0:
        data_rows = data_rows[:import_count]

    columns, data_rows = apply_column_filter(columns, data_rows, fields.get("columnFilter", ""))
    return TabularData(columns=columns, rows=data_rows, sheets=sheets or [], selected_sheet=selected_sheet)


def apply_column_filter(columns: list[str], rows: list[list[str]], filter_value: str) -> tuple[list[str], list[list[str]]]:
    filters = split_values(filter_value)
    if not filters:
        return columns, rows
    indexes: list[int] = []
    for item in filters:
        if item.isdigit():
            index = int(item) - 1
            if 0 <= index < len(columns):
                indexes.append(index)
        elif item in columns:
            indexes.append(columns.index(item))
    if not indexes:
        raise ValueError("指定导入列没有匹配到任何字段。")
    return [columns[i] for i in indexes], [[row[i] if i < len(row) else "" for i in indexes] for row in rows]


def read_tabular_file(path: Path, fields: dict[str, str]) -> TabularData:
    suffix = path.suffix.lower()
    if suffix in {".csv", ".txt"}:
        rows = read_csv_rows(path, fields.get("encoding", "auto"), fields.get("delimiter", ""), decode_escaped(fields.get("lineDelimiter", "")))
        return make_tabular(rows, fields)
    if suffix in {".xlsx", ".xlsm"}:
        rows, sheets, selected = read_excel_rows(path, fields)
        return make_tabular(rows, fields, sheets, selected)
    if suffix == ".xls":
        rows, sheets, selected = read_xls_rows(path, fields)
        return make_tabular(rows, fields, sheets, selected)
    if suffix == ".json":
        rows = read_json_rows(path, fields.get("rowTag", ""))
        with_header = dict(fields)
        with_header["hasHeader"] = "true"
        return make_tabular(rows, with_header)
    if suffix == ".xml":
        rows = read_xml_rows(path, fields.get("rowTag", ""))
        with_header = dict(fields)
        with_header["hasHeader"] = "true"
        return make_tabular(rows, with_header)
    if suffix == ".dbf":
        rows = read_dbf_rows(path, fields)
        with_header = dict(fields)
        with_header["hasHeader"] = "true"
        return make_tabular(rows, with_header)
    supported = ", ".join(sorted(SUPPORTED_EXTENSIONS))
    raise ValueError(f"暂不支持 {suffix or '未知'} 文件。当前支持：{supported}")


def read_tabular_tasks(path: Path, fields: dict[str, str]) -> list[TabularData]:
    suffix = path.suffix.lower()
    if fields.get("sheetMode", "specified") != "all" or suffix not in {".xlsx", ".xlsm", ".xls"}:
        return [read_tabular_file(path, fields)]
    if suffix == ".xls":
        names = xlrd.open_workbook(path).sheet_names()
    else:
        workbook = load_workbook(path, read_only=True, data_only=True)
        names = workbook.sheetnames
        workbook.close()
    tasks: list[TabularData] = []
    for name in names:
        per_sheet = dict(fields)
        per_sheet["sheetFilterMode"] = "name"
        per_sheet["sheetName"] = name
        tasks.append(read_tabular_file(path, per_sheet))
    return tasks

def transform_field_name(name: str, fields: dict[str, str], fallback: str) -> str:
    value = name.strip() or fallback
    if parse_bool(fields, "fieldPinyin", False):
        value = to_pinyin_initials(value)
    replace_from = fields.get("fieldReplaceFrom", "")
    replace_to = fields.get("fieldReplaceTo", "_")
    if replace_from == "symbol":
        value = re.sub(r"[^\w\u4e00-\u9fff]+", replace_to, value, flags=re.UNICODE)
    elif replace_from == "space":
        value = value.replace(" ", replace_to)
    field_case = fields.get("fieldCase", "lower")
    if field_case == "upper":
        value = value.upper()
    elif field_case == "lower":
        value = value.lower()
    return sanitize_identifier(value, fallback)


def parse_mapping(raw_value: str | None, source_columns: list[str]) -> list[dict[str, object]]:
    if not raw_value:
        return [
            {"sourceIndex": index, "source": column, "target": column, "enabled": True, "defaultValue": "", "matchKey": index == 0}
            for index, column in enumerate(source_columns)
        ]
    payload = json.loads(raw_value)
    if not isinstance(payload, list):
        raise ValueError("字段映射必须是数组。")
    if not payload:
        return [
            {"sourceIndex": index, "source": column, "target": column, "enabled": True, "defaultValue": "", "matchKey": index == 0}
            for index, column in enumerate(source_columns)
        ]
    mapping: list[dict[str, object]] = []
    for index, item in enumerate(payload):
        if not isinstance(item, dict):
            continue
        source_index = int(item.get("sourceIndex", -1))
        if source_index < 0 or source_index >= len(source_columns):
            continue
        mapping.append(
            {
                "sourceIndex": source_index,
                "source": source_columns[source_index],
                "target": str(item.get("target") or source_columns[source_index]),
                "enabled": bool(item.get("enabled", True)),
                "defaultValue": str(item.get("defaultValue") or ""),
                "matchKey": bool(item.get("matchKey", index == 0)),
            }
        )
    return mapping


def column_indexes(columns: list[str], selector: str) -> list[int]:
    indexes: list[int] = []
    lower_map = {column.lower(): index for index, column in enumerate(columns)}
    for item in split_values(selector):
        if item.isdigit():
            index = int(item) - 1
            if 0 <= index < len(columns):
                indexes.append(index)
        elif item in columns:
            indexes.append(columns.index(item))
        elif item.lower() in lower_map:
            indexes.append(lower_map[item.lower()])
    return indexes


def parse_date_column_formats(columns: list[str], selector: str) -> dict[int, str]:
    result: dict[int, str] = {}
    lower_map = {column.lower(): index for index, column in enumerate(columns)}
    for item in split_values(selector):
        if ":" in item:
            key, fmt = item.split(":", 1)
        elif "=" in item:
            key, fmt = item.split("=", 1)
        else:
            continue
        key = key.strip()
        fmt = fmt.strip()
        index = None
        if key.isdigit():
            candidate = int(key) - 1
            if 0 <= candidate < len(columns):
                index = candidate
        elif key in columns:
            index = columns.index(key)
        elif key.lower() in lower_map:
            index = lower_map[key.lower()]
        if index is not None and fmt:
            result[index] = fmt
    return result


def apply_cleaning(columns: list[str], rows: list[list[str]], fields: dict[str, str]) -> tuple[list[str], list[list[str]], int]:
    trim_values = parse_bool(fields, "trimValues", True)
    empty_as_null = parse_bool(fields, "emptyAsNull", False)
    zero_for_number = parse_bool(fields, "zeroForNumber", False)
    replace_blank_with = fields.get("replaceBlankWith", "")
    remove_text = fields.get("removeText", "")
    replace_text_from = fields.get("replaceTextFrom", "")
    replace_text_to = fields.get("replaceTextTo", "")
    blank_values = set(split_values(fields.get("blankCellValues", "")))
    fill_down_indexes = column_indexes(columns, fields.get("fillDownColumns", ""))
    dedupe_indexes = column_indexes(columns, fields.get("dedupeColumns", ""))
    date_formats = parse_date_column_formats(columns, fields.get("dateColumns", ""))

    cleaned: list[list[str | None]] = []
    previous: dict[int, str] = {}
    skipped = 0
    seen: set[tuple[str, ...]] = set()

    for row in rows:
        values: list[str | None] = []
        for index, raw in enumerate(row):
            value = raw or ""
            if trim_values:
                value = value.strip()
            if value in blank_values:
                value = ""
            if remove_text:
                value = value.replace(remove_text, "")
            if replace_text_from:
                value = value.replace(replace_text_from, replace_text_to)
            if value == "" and index in fill_down_indexes and index in previous:
                value = previous[index]
            if value:
                previous[index] = value
            if value == "" and replace_blank_with:
                value = replace_blank_with
            if value == "" and zero_for_number:
                value = "0"
            if value and index in date_formats:
                try:
                    value = dt.datetime.strptime(value, date_formats[index]).isoformat(sep=" ", timespec="seconds")
                except ValueError as exc:
                    raise ValueError(f"日期列 {columns[index]} 的值 {value} 不符合格式 {date_formats[index]}") from exc
            values.append(None if empty_as_null and value == "" else value)

        if dedupe_indexes:
            key = tuple("" if values[i] is None else str(values[i]) for i in dedupe_indexes)
            if key in seen:
                skipped += 1
                continue
            seen.add(key)
        cleaned.append(values)

    return columns, cleaned, skipped


def build_target_data(tabular: TabularData, fields: dict[str, str], file_name: str) -> tuple[list[str], list[list[object]], list[str], int]:
    mapping = [item for item in parse_mapping(fields.get("mapping"), tabular.columns) if item.get("enabled")]
    if not mapping:
        raise ValueError("至少需要启用一个字段。")

    raw_columns = [str(item["target"]) for item in mapping]
    raw_rows = []
    for row in tabular.rows:
        values = []
        for item in mapping:
            source_index = int(item["sourceIndex"])
            value = row[source_index] if source_index < len(row) else ""
            if value == "" and (parse_bool(fields, "defaultForEmpty", False) or item.get("defaultValue")):
                value = str(item.get("defaultValue") or "")
            values.append(value)
        raw_rows.append(values)

    transformed_columns = unique_names([transform_field_name(name, fields, f"column_{i + 1}") for i, name in enumerate(raw_columns)])
    transformed_columns, cleaned_rows, skipped = apply_cleaning(transformed_columns, raw_rows, fields)

    match_keys = [
        transformed_columns[index]
        for index, item in enumerate(mapping)
        if index < len(transformed_columns) and item.get("matchKey")
    ]
    if not match_keys and transformed_columns:
        match_keys = [transformed_columns[0]]

    final_columns = list(transformed_columns)
    final_rows: list[list[object]] = [list(row) for row in cleaned_rows]

    auto_pk_field = fields.get("autoPkField", "").strip()
    if auto_pk_field:
        column = transform_field_name(auto_pk_field, fields, "id")
        final_columns.insert(0, column)
        for index, row in enumerate(final_rows, start=1):
            row.insert(0, index)

    extras: list[tuple[str, object]] = []
    if fields.get("importTimeField", "").strip():
        extras.append((transform_field_name(fields["importTimeField"], fields, "imported_at"), dt.datetime.now().isoformat(sep=" ", timespec="seconds")))
    if fields.get("sheetNameField", "").strip():
        extras.append((transform_field_name(fields["sheetNameField"], fields, "sheet_name"), tabular.selected_sheet or Path(file_name).stem))
    if fields.get("fixedValueField", "").strip():
        extras.append((transform_field_name(fields["fixedValueField"], fields, "fixed_value"), fields.get("fixedValue", "")))

    for column, value in extras:
        final_columns.append(column)
        for row in final_rows:
            row.append(value)

    return final_columns, final_rows, match_keys, skipped


def normalize_target_name(uploaded: UploadedFile, tabular: TabularData, fields: dict[str, str]) -> str:
    if fields.get("tableName"):
        base = fields["tableName"]
    elif fields.get("tableNameRule") == "sheet" and tabular.selected_sheet:
        base = tabular.selected_sheet
    else:
        base = Path(uploaded.filename).stem

    regex = fields.get("tableRegex", "").strip()
    if regex:
        match = re.search(regex, base)
        if match:
            base = match.group(1) if match.groups() else match.group(0)

    if parse_bool(fields, "symbolToUnderscore", False):
        base = re.sub(r"[^\w\u4e00-\u9fff]+", "_", base, flags=re.UNICODE)
    if parse_bool(fields, "tablePinyin", False):
        base = to_pinyin_initials(base)

    value = f"{fields.get('tablePrefix', '')}{base}{fields.get('tableSuffix', '')}"
    target_case = fields.get("tableCase", "lower")
    if target_case == "upper":
        value = value.upper()
    elif target_case == "lower":
        value = value.lower()
    return sanitize_identifier(value, "import_table")


def quote_identifier(name: str) -> str:
    return '"' + name.replace('"', '""') + '"'


def existing_columns(conn: sqlite3.Connection, table_name: str) -> list[str]:
    rows = conn.execute(f"pragma table_info({quote_identifier(table_name)})").fetchall()
    return [row["name"] for row in rows]


def table_exists(conn: sqlite3.Connection, table_name: str) -> bool:
    return (
        conn.execute(
            "select 1 from sqlite_master where type = 'table' and name = ? limit 1",
            (table_name,),
        ).fetchone()
        is not None
    )


def target_table_exists(conn, table_name: str, fields: dict[str, str]) -> bool:
    if target_db_type(fields) == "mysql":
        with conn.cursor() as cursor:
            cursor.execute(
                "select 1 from information_schema.tables where table_schema = database() and table_name = %s limit 1",
                (table_name,),
            )
            return cursor.fetchone() is not None
    return table_exists(conn, table_name)


def target_existing_columns(conn, table_name: str, fields: dict[str, str]) -> list[str]:
    if target_db_type(fields) == "mysql":
        with conn.cursor() as cursor:
            cursor.execute(
                """
                select column_name from information_schema.columns
                where table_schema = database() and table_name = %s
                order by ordinal_position
                """,
                (table_name,),
            )
            return [row[0] for row in cursor.fetchall()]
    return existing_columns(conn, table_name)


def infer_column_types(columns: list[str], rows: list[list[object]], fields: dict[str, str]) -> dict[str, str]:
    db_type = target_db_type(fields)
    text_type = "text" if db_type == "sqlite" else "text"
    int_type = "integer" if db_type == "sqlite" else "bigint"
    real_type = "real" if db_type == "sqlite" else "double"
    if fields.get("typeMode", "auto") == "text":
        return {column: text_type for column in columns}
    types: dict[str, str] = {}
    for index, column in enumerate(columns):
        values = [row[index] for row in rows if index < len(row) and row[index] not in (None, "")]
        if values and all(re.fullmatch(r"[-+]?\d+", str(value)) for value in values):
            types[column] = int_type
        elif values and all(re.fullmatch(r"[-+]?(\d+(\.\d*)?|\.\d+)", str(value)) for value in values):
            types[column] = real_type
        else:
            types[column] = text_type
    return types


def target_create_or_expand_table(conn, table_name: str, columns: list[str], rows: list[list[object]], rebuild: bool, allow_expand: bool, fields: dict[str, str]) -> None:
    table = db_quote(table_name, fields)
    column_types = infer_column_types(columns, rows, fields)
    if rebuild:
        with conn.cursor() if target_db_type(fields) == "mysql" else nullcontext(conn) as cursor:
            cursor.execute(f"drop table if exists {table}")
    definitions = ", ".join(f"{db_quote(column, fields)} {column_types[column]}" for column in columns)
    sql = f"create table if not exists {table} ({definitions})"
    if target_db_type(fields) == "mysql":
        with conn.cursor() as cursor:
            cursor.execute(sql)
    else:
        conn.execute(sql)
    existing = target_existing_columns(conn, table_name, fields)
    missing = [column for column in columns if column not in existing]
    if missing and not allow_expand:
        raise ValueError(f"目标表缺少字段：{', '.join(missing)}")
    for column in missing:
        sql = f"alter table {table} add column {db_quote(column, fields)} {column_types[column]}"
        if target_db_type(fields) == "mysql":
            with conn.cursor() as cursor:
                cursor.execute(sql)
        else:
            conn.execute(sql)


def target_insert_rows(conn, table_name: str, columns: list[str], rows: list[list[object]], fields: dict[str, str], progress=None) -> int:
    if not rows:
        return 0
    table = db_quote(table_name, fields)
    quoted_columns = ", ".join(db_quote(column, fields) for column in columns)
    placeholders = ", ".join(db_placeholder(fields) for _ in columns)
    sql = f"insert into {table} ({quoted_columns}) values ({placeholders})"
    batch_size = max(parse_int(fields, "batchRows", 0), 0)
    batches = [rows] if batch_size <= 0 else [rows[index : index + batch_size] for index in range(0, len(rows), batch_size)]
    total = 0
    if target_db_type(fields) == "mysql":
        with conn.cursor() as cursor:
            for batch in batches:
                cursor.executemany(sql, [row[: len(columns)] for row in batch])
                total += len(batch)
                if progress:
                    progress(total)
                if fields.get("commitMode") == "batch":
                    conn.commit()
    else:
        for batch in batches:
            conn.executemany(sql, [row[: len(columns)] for row in batch])
            total += len(batch)
            if progress:
                progress(total)
            if fields.get("commitMode") == "batch":
                conn.commit()
    return total


def target_insert_rows_parallel(table_name: str, columns: list[str], rows: list[list[object]], fields: dict[str, str], progress=None) -> int:
    if not rows:
        return 0
    workers = max(parse_int(fields, "parallelWorkers", 4), 2)
    batch_size = max(parse_int(fields, "batchRows", 1000), 1)
    chunks = [rows[index : index + batch_size] for index in range(0, len(rows), batch_size)]
    completed = 0

    def write_chunk(chunk: list[list[object]]) -> int:
        conn = connect_target_db(fields)
        try:
            target_insert_rows(conn, table_name, columns, chunk, {**fields, "writeMode": "fast"}, None)
            if target_db_type(fields) != "mysql" or fields.get("commitMode") != "auto":
                conn.commit()
            return len(chunk)
        finally:
            conn.close()

    with ThreadPoolExecutor(max_workers=workers) as executor:
        for count in executor.map(write_chunk, chunks):
            completed += count
            if progress:
                progress(completed)
    return completed


def mysql_load_rows(conn, table_name: str, columns: list[str], rows: list[list[object]], fields: dict[str, str], progress=None) -> int:
    if target_db_type(fields) != "mysql" or not rows:
        return 0
    fd, temp_name = tempfile.mkstemp(prefix="codex_load_", suffix=".tsv")
    os.close(fd)
    temp_path = Path(temp_name)
    try:
        with temp_path.open("w", encoding="utf-8", newline="") as file:
            writer = csv.writer(file, delimiter="\t", lineterminator="\n", quoting=csv.QUOTE_MINIMAL)
            for row in rows:
                writer.writerow(["\\N" if value is None else value for value in row[: len(columns)]])
        sql = (
            f"load data local infile {db_placeholder(fields)} into table {db_quote(table_name, fields)} "
            "character set utf8mb4 fields terminated by '\\t' optionally enclosed by '\"' "
            "lines terminated by '\\n' "
            f"({', '.join(db_quote(column, fields) for column in columns)})"
        )
        with conn.cursor() as cursor:
            cursor.execute(sql, (str(temp_path).replace("\\", "/"),))
        if progress:
            progress(len(rows))
        return len(rows)
    finally:
        temp_path.unlink(missing_ok=True)


def target_update_rows(conn, table_name: str, columns: list[str], rows: list[list[object]], match_keys: list[str], fields: dict[str, str]) -> tuple[int, int]:
    if not match_keys:
        raise ValueError("更新模式需要至少一个匹配键。")
    key_indexes = [columns.index(key) for key in match_keys if key in columns]
    if not key_indexes:
        raise ValueError("匹配键不在导入字段中。")
    update_columns = [column for column in columns if column not in match_keys]
    inserted = 0
    updated = 0
    table = db_quote(table_name, fields)
    for row in rows:
        where = " and ".join(f"{db_quote(columns[index], fields)} <=> {db_placeholder(fields)}" if target_db_type(fields) == "mysql" else f"{db_quote(columns[index], fields)} is {db_placeholder(fields)}" for index in key_indexes)
        key_values = [row[index] for index in key_indexes]
        select_sql = f"select 1 from {table} where {where} limit 1"
        if target_db_type(fields) == "mysql":
            with conn.cursor() as cursor:
                cursor.execute(select_sql, key_values)
                exists = cursor.fetchone()
        else:
            exists = conn.execute(select_sql, key_values).fetchone()
        if exists:
            if update_columns:
                assignments = ", ".join(f"{db_quote(column, fields)} = {db_placeholder(fields)}" for column in update_columns)
                values = [row[columns.index(column)] for column in update_columns] + key_values
                update_sql = f"update {table} set {assignments} where {where}"
                if target_db_type(fields) == "mysql":
                    with conn.cursor() as cursor:
                        cursor.execute(update_sql, values)
                else:
                    conn.execute(update_sql, values)
            updated += 1
        else:
            target_insert_rows(conn, table_name, columns, [row], fields)
            inserted += 1
    return inserted, updated


def target_execute_sql_batch(conn, sql_text: str, label: str, fields: dict[str, str]) -> None:
    sql_text = (sql_text or "").strip()
    if not sql_text:
        return
    try:
        if target_db_type(fields) == "mysql":
            with conn.cursor() as cursor:
                for statement in [part.strip() for part in sql_text.split(";") if part.strip()]:
                    cursor.execute(statement)
        else:
            conn.executescript(sql_text)
    except Exception as exc:
        raise ValueError(f"{label} 执行失败：{exc}") from exc


def target_export_query_to_excel(conn, sql_text: str, output_name: str, fields: dict[str, str]) -> str:
    sql_text = (sql_text or "").strip()
    if not sql_text:
        return ""
    requested = Path(output_name.strip() or f"query_result_{int(time.time())}.xlsx").name
    output = export_target_path(Path(requested).stem, "xlsx", fields)
    workbook = Workbook()
    sheet = workbook.active
    if target_db_type(fields) == "mysql":
        with conn.cursor() as cursor:
            cursor.execute(sql_text)
            rows = cursor.fetchall()
            columns = [desc[0] for desc in cursor.description or []]
        if columns:
            sheet.append(columns)
            for row in rows:
                sheet.append(list(row))
    else:
        rows = conn.execute(sql_text).fetchall()
        if rows:
            columns = rows[0].keys()
            sheet.append(list(columns))
            for row in rows:
                sheet.append([row[column] for column in columns])
    workbook.save(output)
    return str(output)


def create_or_expand_table(conn: sqlite3.Connection, table_name: str, columns: list[str], rebuild: bool, allow_expand: bool) -> None:
    table = quote_identifier(table_name)
    if rebuild:
        conn.execute(f"drop table if exists {table}")
    definitions = ", ".join(f"{quote_identifier(column)} text" for column in columns)
    conn.execute(f"create table if not exists {table} ({definitions})")
    existing = existing_columns(conn, table_name)
    missing = [column for column in columns if column not in existing]
    if missing and not allow_expand:
        raise ValueError(f"目标表缺少字段：{', '.join(missing)}")
    for column in missing:
        conn.execute(f"alter table {table} add column {quote_identifier(column)} text")


def insert_rows(conn: sqlite3.Connection, table_name: str, columns: list[str], rows: list[list[object]]) -> int:
    if not rows:
        return 0
    table = quote_identifier(table_name)
    quoted_columns = ", ".join(quote_identifier(column) for column in columns)
    placeholders = ", ".join("?" for _ in columns)
    sql = f"insert into {table} ({quoted_columns}) values ({placeholders})"
    conn.executemany(sql, [row[: len(columns)] for row in rows])
    return len(rows)


def update_rows(conn: sqlite3.Connection, table_name: str, columns: list[str], rows: list[list[object]], match_keys: list[str]) -> tuple[int, int]:
    if not match_keys:
        raise ValueError("更新模式需要至少一个匹配键。")
    key_indexes = [columns.index(key) for key in match_keys if key in columns]
    if not key_indexes:
        raise ValueError("匹配键不在导入字段中。")

    update_columns = [column for column in columns if column not in match_keys]
    updated = 0
    inserted = 0
    table = quote_identifier(table_name)

    for row in rows:
        where = " and ".join(f"{quote_identifier(columns[index])} is ?" for index in key_indexes)
        key_values = [row[index] for index in key_indexes]
        exists = conn.execute(f"select 1 from {table} where {where} limit 1", key_values).fetchone()
        if exists:
            if update_columns:
                assignments = ", ".join(f"{quote_identifier(column)} = ?" for column in update_columns)
                values = [row[columns.index(column)] for column in update_columns] + key_values
                conn.execute(f"update {table} set {assignments} where {where}", values)
            updated += 1
        else:
            insert_rows(conn, table_name, columns, [row])
            inserted += 1
    return inserted, updated


def execute_sql_batch(conn: sqlite3.Connection, sql_text: str, label: str) -> None:
    sql_text = (sql_text or "").strip()
    if not sql_text:
        return
    try:
        conn.executescript(sql_text)
    except sqlite3.Error as exc:
        raise ValueError(f"{label} 执行失败：{exc}") from exc


def export_query_to_excel(conn: sqlite3.Connection, sql_text: str, output_name: str) -> str:
    sql_text = (sql_text or "").strip()
    if not sql_text:
        return ""
    output = EXPORTS / (output_name.strip() or f"query_result_{int(time.time())}.xlsx")
    if output.suffix.lower() != ".xlsx":
        output = output.with_suffix(".xlsx")
    rows = conn.execute(sql_text).fetchall()
    workbook = Workbook()
    sheet = workbook.active
    if rows:
        columns = rows[0].keys()
        sheet.append(list(columns))
        for row in rows:
            sheet.append([row[column] for column in columns])
    workbook.save(output)
    return str(output)


def safe_file_stem(value: str, fallback: str = "export") -> str:
    stem = re.sub(r'[<>:"/\\|?*\x00-\x1f]+', "_", (value or "").strip())
    stem = re.sub(r"\s+", " ", stem).strip(" ._")
    return (stem or fallback)[:120]


def safe_sheet_name(value: str, fallback: str = "Sheet1") -> str:
    name = re.sub(r"[:\\/?*\[\]]+", "_", (value or "").strip())
    return (name or fallback)[:31]


def export_sources(fields: dict[str, str]) -> list[dict[str, object]]:
    conn = connect_target_db(fields)
    try:
        if target_db_type(fields) == "mysql":
            with conn.cursor() as cursor:
                cursor.execute(
                    """
                    select table_name, table_type, table_comment, table_rows
                    from information_schema.tables
                    where table_schema = database()
                    order by table_name
                    """
                )
                rows = cursor.fetchall()
            return [
                {
                    "name": row[0],
                    "type": row[1],
                    "comment": row[2] or "",
                    "rows": row[3] or 0,
                }
                for row in rows
            ]
        rows = conn.execute(
            """
            select name, type
            from sqlite_master
            where type in ('table', 'view') and name not like 'sqlite_%' and name not like '\\_%' escape '\\'
            order by name
            """
        ).fetchall()
        result = []
        for row in rows:
            count = 0
            try:
                count = conn.execute(f"select count(*) from {db_quote(row['name'], fields)}").fetchone()[0]
            except Exception:
                count = 0
            result.append({"name": row["name"], "type": row["type"], "comment": "", "rows": count})
        return result
    finally:
        conn.close()


def export_split_list(value: object) -> list[str]:
    if isinstance(value, list):
        return [str(item).strip() for item in value if str(item).strip()]
    return split_values(str(value or ""))


def export_field_list(value: object) -> list[str]:
    return export_split_list(value)


def export_query_from_item(item: dict[str, object], fields: dict[str, str]) -> tuple[str, str]:
    item_type = str(item.get("type") or "table")
    if item_type == "query":
        sql = str(item.get("sql") or "").strip()
        if not sql:
            raise ValueError("查询 SQL 不能为空。")
        return sql, str(item.get("name") or "query")
    table_name = str(item.get("table") or item.get("name") or "").strip()
    if not table_name:
        raise ValueError("请选择要导出的表。")
    selected_fields = export_field_list(fields.get("exportFields", ""))
    columns_sql = ", ".join(db_quote(column, fields) for column in selected_fields) if selected_fields else "*"
    sql = f"select {columns_sql} from {db_quote(table_name, fields)}"
    where = str(fields.get("whereClause") or "").strip()
    if where:
        sql += " where " + re.sub(r"^\s*where\s+", "", where, flags=re.I)
    return sql, table_name


def fetch_export_rows(conn, sql: str, fields: dict[str, str], limit: int = 0) -> tuple[list[str], list[list[object]]]:
    query = sql.strip().rstrip(";")
    if limit > 0:
        if target_db_type(fields) == "mysql":
            query = f"select * from ({query}) export_preview limit {limit}"
        else:
            query = f"select * from ({query}) limit {limit}"
    if target_db_type(fields) == "mysql":
        with conn.cursor() as cursor:
            cursor.execute(query)
            columns = [desc[0] for desc in cursor.description or []]
            rows = [list(row) for row in cursor.fetchall()]
        return columns, rows
    cursor = conn.execute(query)
    columns = [desc[0] for desc in cursor.description or []]
    return columns, [list(row) for row in cursor.fetchall()]


def export_row_batches(conn, sql: str, fields: dict[str, str], fetch_size: int = EXPORT_FETCH_SIZE) -> Iterator[tuple[list[str], list[list[object]]]]:
    query = sql.strip().rstrip(";")
    if target_db_type(fields) == "mysql":
        cursor = conn.cursor(pymysql.cursors.SSCursor)
        try:
            cursor.execute(query)
            columns = [desc[0] for desc in cursor.description or []]
            while True:
                batch = cursor.fetchmany(fetch_size)
                if not batch:
                    break
                yield columns, [list(row) for row in batch]
        finally:
            cursor.close()
        return

    cursor = conn.execute(query)
    columns = [desc[0] for desc in cursor.description or []]
    while True:
        batch = cursor.fetchmany(fetch_size)
        if not batch:
            break
        yield columns, [list(row) for row in batch]


def add_export_time_column(columns: list[str], rows: list[list[object]], field_name: str) -> tuple[list[str], list[list[object]]]:
    field_name = (field_name or "").strip()
    if not field_name:
        return columns, rows
    now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    return columns + [field_name], [row + [now] for row in rows]


def add_export_time_to_batch(rows: list[list[object]], field_name: str, now: str) -> list[list[object]]:
    if not field_name:
        return rows
    return [row + [now] for row in rows]


def apply_export_batch_limit(rows: list[list[object]], fields: dict[str, str]) -> list[list[object]]:
    batch_rows = int(fields.get("batchRows") or 0)
    if batch_rows > 0 and parse_bool(fields, "splitByBatch", False):
        return rows[:batch_rows]
    return rows


def rows_to_dicts(columns: list[str], rows: list[list[object]]) -> list[dict[str, object]]:
    return [{column: row[index] if index < len(row) else None for index, column in enumerate(columns)} for row in rows]


def row_to_dict(columns: list[str], row: list[object]) -> dict[str, object]:
    return {column: row[index] if index < len(row) else None for index, column in enumerate(columns)}


def export_target_path(base_name: str, extension: str, fields: dict[str, str]) -> Path:
    extension = extension.lower().lstrip(".") or "xlsx"
    if extension == "xls":
        raise ValueError("当前版本不支持 .xls 导出，请选择 .xlsx。")
    if extension == "dbf":
        raise ValueError("当前版本暂不支持 DBF 导出。")
    prefix = safe_file_stem(str(fields.get("filePrefix") or ""), "")
    suffix = safe_file_stem(str(fields.get("fileSuffix") or ""), "")
    name = safe_file_stem(f"{prefix}{base_name}{suffix}", "export")
    return EXPORTS / f"{name}.{extension}"


def write_rows_to_sheet(sheet, columns: list[str], rows: list[list[object]], fields: dict[str, str]) -> None:
    header_mode = str(fields.get("headerMode") or "field").lower()
    include_header = header_mode != "none"
    if include_header:
        sheet.append(columns)
    for row in rows:
        sheet.append(row)

    row_height = float(fields.get("rowHeight") or 0)
    if row_height > 0:
        for row_idx in range(1, sheet.max_row + 1):
            sheet.row_dimensions[row_idx].height = row_height
    col_width = float(fields.get("columnWidth") or 0)
    if col_width > 0:
        for col_idx in range(1, sheet.max_column + 1):
            sheet.column_dimensions[get_column_letter(col_idx)].width = col_width

    font_name = str(fields.get("fontName") or "").strip()
    font_size = float(fields.get("fontSize") or 0)
    font = Font(name=font_name or None, size=font_size or None)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    add_border = parse_bool(fields, "addBorder", False)
    for row in sheet.iter_rows():
        for cell in row:
            if font_name or font_size:
                cell.font = font
            if add_border:
                cell.border = border

    if parse_bool(fields, "lockHeader", False) and include_header:
        sheet.protection.sheet = True
        for cell in sheet[1]:
            cell.protection = Protection(locked=True)
    locked_columns = export_split_list(fields.get("lockedColumns", ""))
    if locked_columns:
        sheet.protection.sheet = True
        column_indexes = {name: index + 1 for index, name in enumerate(columns)}
        for name in locked_columns:
            col_idx = column_indexes.get(name)
            if col_idx:
                for row_idx in range(1, sheet.max_row + 1):
                    sheet.cell(row=row_idx, column=col_idx).protection = Protection(locked=True)


def write_export_file(path: Path, columns: list[str], rows: list[list[object]], fields: dict[str, str], sheet_name: str) -> None:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    extension = path.suffix.lower()
    if extension == ".xlsx":
        mode = str(fields.get("exportMode") or "workbook")
        if path.exists() and mode in {"sheet", "data"}:
            workbook = load_workbook(path)
            if sheet_name in workbook.sheetnames:
                del workbook[sheet_name]
            sheet = workbook.create_sheet(sheet_name)
        else:
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = sheet_name
        write_rows_to_sheet(sheet, columns, rows, fields)
        workbook.save(path)
        return
    if extension in {".csv", ".txt"}:
        delimiter = decode_escaped(fields.get("delimiter") or ("," if extension == ".csv" else "\t"))
        line_delimiter = decode_escaped(fields.get("lineDelimiter") or "\n")
        encoding = fields.get("encoding") or "utf-8"
        with path.open("w", encoding=encoding, newline="") as handle:
            writer = csv.writer(handle, delimiter=delimiter, lineterminator=line_delimiter)
            if str(fields.get("headerMode") or "field") != "none":
                writer.writerow(columns)
            writer.writerows(rows)
        return
    if extension == ".json":
        path.write_text(json.dumps(rows_to_dicts(columns, rows), ensure_ascii=False, indent=2, default=cell_to_text), encoding="utf-8")
        return
    if extension == ".xml":
        root = ET.Element("rows")
        for row in rows_to_dicts(columns, rows):
            node = ET.SubElement(root, "row")
            for key, value in row.items():
                child = ET.SubElement(node, sanitize_identifier(key, "field"))
                child.text = cell_to_text(value)
        ET.ElementTree(root).write(path, encoding="utf-8", xml_declaration=True)
        return
    raise ValueError(f"不支持的导出格式：{extension}")


def styled_write_only_row(sheet, values: list[object], fields: dict[str, str], header: bool = False) -> None:
    font_name = str(fields.get("fontName") or "").strip()
    font_size = float(fields.get("fontSize") or 0)
    add_border = parse_bool(fields, "addBorder", False)
    if not (font_name or font_size or add_border or header):
        sheet.append(values)
        return

    font = Font(name=font_name or None, size=font_size or None, bold=header)
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    cells = []
    for value in values:
        cell = WriteOnlyCell(sheet, value=value)
        if font_name or font_size or header:
            cell.font = font
        if add_border:
            cell.border = border
        cells.append(cell)
    sheet.append(cells)


def write_export_file_streaming(
    path: Path,
    columns: list[str],
    batches: Iterable[list[list[object]]],
    fields: dict[str, str],
    sheet_name: str,
) -> int:
    EXPORTS.mkdir(parents=True, exist_ok=True)
    extension = path.suffix.lower()
    header_mode = str(fields.get("headerMode") or "field").lower()
    include_header = header_mode != "none"
    rows_written = 0

    if extension == ".xlsx":
        workbook = Workbook(write_only=True)
        sheet = workbook.create_sheet(sheet_name)
        row_height = float(fields.get("rowHeight") or 0)
        if row_height > 0:
            sheet.sheet_format.defaultRowHeight = row_height
        col_width = float(fields.get("columnWidth") or 0)
        if col_width > 0:
            for col_idx in range(1, len(columns) + 1):
                sheet.column_dimensions[get_column_letter(col_idx)].width = col_width
        if include_header:
            styled_write_only_row(sheet, columns, fields, header=True)
        for batch in batches:
            for row in batch:
                styled_write_only_row(sheet, row, fields)
                rows_written += 1
        workbook.save(path)
        return rows_written

    if extension in {".csv", ".txt"}:
        delimiter = decode_escaped(fields.get("delimiter") or ("," if extension == ".csv" else "\t"))
        line_delimiter = decode_escaped(fields.get("lineDelimiter") or "\n")
        encoding = fields.get("encoding") or "utf-8"
        with path.open("w", encoding=encoding, newline="") as handle:
            writer = csv.writer(handle, delimiter=delimiter, lineterminator=line_delimiter)
            if include_header:
                writer.writerow(columns)
            for batch in batches:
                writer.writerows(batch)
                rows_written += len(batch)
        return rows_written

    if extension == ".json":
        with path.open("w", encoding="utf-8") as handle:
            handle.write("[\n")
            first = True
            for batch in batches:
                for row in batch:
                    if not first:
                        handle.write(",\n")
                    handle.write(json.dumps(row_to_dict(columns, row), ensure_ascii=False, default=cell_to_text))
                    first = False
                    rows_written += 1
            handle.write("\n]\n")
        return rows_written

    if extension == ".xml":
        with path.open("w", encoding="utf-8") as handle:
            handle.write('<?xml version="1.0" encoding="utf-8"?>\n<rows>\n')
            safe_columns = [sanitize_identifier(column, "field") for column in columns]
            for batch in batches:
                for row in batch:
                    handle.write("  <row>\n")
                    for index, column in enumerate(safe_columns):
                        value = cell_to_text(row[index] if index < len(row) else "")
                        handle.write(f"    <{column}>{xml_escape(value)}</{column}>\n")
                    handle.write("  </row>\n")
                    rows_written += 1
            handle.write("</rows>\n")
        return rows_written

    raise ValueError(f"不支持的导出格式：{extension}")


def group_rows_by_field(columns: list[str], rows: list[list[object]], split_field: str) -> dict[str, list[list[object]]]:
    if not split_field:
        return {"": rows}
    if split_field not in columns:
        raise ValueError(f"拆分字段不存在：{split_field}")
    index = columns.index(split_field)
    groups: dict[str, list[list[object]]] = {}
    for row in rows:
        key = safe_file_stem(cell_to_text(row[index] if index < len(row) else ""), "empty")
        groups.setdefault(key, []).append(row)
    return groups


def run_export_job(payload: dict[str, object]) -> dict[str, object]:
    fields = {key: str(value) for key, value in payload.items() if not isinstance(value, (list, dict))}
    if fields.get("targetDbType") != "sqlite" and not fields.get("connectionId"):
        raise ValueError("请选择数据库连接。")
    items = payload.get("items")
    if not isinstance(items, list) or not items:
        single = {"type": payload.get("sourceType") or "table", "name": payload.get("table") or "query", "table": payload.get("table"), "sql": payload.get("sql")}
        items = [single]
    extension = str(payload.get("extension") or fields.get("extension") or "xlsx").lower().lstrip(".")
    if extension not in {"xlsx", "csv", "txt", "json", "xml"}:
        raise ValueError("当前仅支持 xlsx、csv、txt、json、xml 导出。")

    if parse_bool(fields, "clearLogBeforeExport", False):
        (EXPORTS / "export.log").write_text("", encoding="utf-8")

    conn = connect_target_db(fields)
    started = time.time()
    written_files: list[str] = []
    total_rows = 0
    try:
        target_execute_sql_batch(conn, str(payload.get("beforeSql") or ""), "导出开始前 SQL", fields)
        split_field = str(fields.get("splitField") or "").strip()
        can_stream = not split_field and str(fields.get("exportMode") or "workbook") == "workbook"

        if can_stream and extension == "xlsx" and len(items) > 1:
            workbook = Workbook(write_only=True)
            for item in items:
                if not isinstance(item, dict):
                    continue
                sql, source_name = export_query_from_item(item, fields)
                sheet_name = safe_sheet_name(str(fields.get("sheetName") or source_name or "Sheet1"))
                sheet = workbook.create_sheet(sheet_name)
                row_count = 0
                export_time_field = str(fields.get("exportTimeField") or "").strip()
                export_time_value = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                batch_limit = int(fields.get("batchRows") or 0) if parse_bool(fields, "splitByBatch", False) else 0
                for columns, batch in export_row_batches(conn, sql, fields):
                    if export_time_field and export_time_field not in columns:
                        columns = columns + [export_time_field]
                    if row_count == 0 and str(fields.get("headerMode") or "field") != "none":
                        styled_write_only_row(sheet, columns, fields, header=True)
                    batch = add_export_time_to_batch(batch, export_time_field, export_time_value)
                    if batch_limit:
                        batch = batch[: max(0, batch_limit - row_count)]
                    for row in batch:
                        styled_write_only_row(sheet, row, fields)
                    row_count += len(batch)
                    if batch_limit and row_count >= batch_limit:
                        break
                if parse_bool(fields, "skipEmptyTable", False) and row_count == 0:
                    continue
                total_rows += row_count
            path = export_target_path(str(fields.get("outputName") or "export"), "xlsx", fields)
            workbook.save(path)
            written_files.append(str(path))
        else:
            for item in items:
                if not isinstance(item, dict):
                    continue
                sql, source_name = export_query_from_item(item, fields)
                sheet_name = safe_sheet_name(str(fields.get("sheetName") or source_name or "Sheet1"))
                base_name = str(fields.get("outputName") or source_name)
                path = export_target_path(base_name, extension, fields)

                if can_stream:
                    first_columns: list[str] = []
                    export_time_field = str(fields.get("exportTimeField") or "").strip()
                    export_time_value = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    batch_limit = int(fields.get("batchRows") or 0) if parse_bool(fields, "splitByBatch", False) else 0
                    rows_seen = 0

                    def streaming_batches() -> Iterator[list[list[object]]]:
                        nonlocal first_columns, rows_seen
                        for columns, batch in export_row_batches(conn, sql, fields):
                            if export_time_field and export_time_field not in columns:
                                columns = columns + [export_time_field]
                            if not first_columns:
                                first_columns = columns
                            batch = add_export_time_to_batch(batch, export_time_field, export_time_value)
                            if batch_limit:
                                batch = batch[: max(0, batch_limit - rows_seen)]
                            rows_seen += len(batch)
                            if batch:
                                yield batch
                            if batch_limit and rows_seen >= batch_limit:
                                break

                    buffered_batches = streaming_batches()
                    try:
                        first_batch = next(buffered_batches)
                    except StopIteration:
                        if parse_bool(fields, "skipEmptyTable", False):
                            continue
                        first_columns = []
                        first_batch = []
                    row_count = write_export_file_streaming(path, first_columns, chain([first_batch], buffered_batches), fields, sheet_name)
                    written_files.append(str(path))
                    total_rows += row_count
                else:
                    columns, rows = fetch_export_rows(conn, sql, fields)
                    columns, rows = add_export_time_column(columns, rows, str(fields.get("exportTimeField") or ""))
                    rows = apply_export_batch_limit(rows, fields)
                    if parse_bool(fields, "skipEmptyTable", False) and not rows:
                        continue
                    groups = group_rows_by_field(columns, rows, split_field)
                    for group_name, group_rows in groups.items():
                        group_base_name = base_name
                        if group_name:
                            group_base_name = f"{group_base_name}_{group_name}"
                        group_path = export_target_path(group_base_name, extension, fields)
                        write_export_file(group_path, columns, group_rows, fields, sheet_name)
                        written_files.append(str(group_path))
                        total_rows += len(group_rows)
        target_execute_sql_batch(conn, str(payload.get("afterSql") or ""), "导出结束后 SQL", fields)
        conn.commit()
    finally:
        conn.close()

    return {
        "files": written_files,
        "rows": total_rows,
        "elapsedMs": int((time.time() - started) * 1000),
    }


def preview_export_job(payload: dict[str, object]) -> dict[str, object]:
    fields = {key: str(value) for key, value in payload.items() if not isinstance(value, (list, dict))}
    items = payload.get("items")
    if isinstance(items, list) and items and isinstance(items[0], dict):
        item = items[0]
    else:
        item = {"type": payload.get("sourceType") or "table", "name": payload.get("table") or "query", "table": payload.get("table"), "sql": payload.get("sql")}
    sql, source_name = export_query_from_item(item, fields)
    conn = connect_target_db(fields)
    try:
        columns, rows = fetch_export_rows(conn, sql, fields, MAX_PREVIEW_ROWS)
    finally:
        conn.close()
    return {"sourceName": source_name, "columns": columns, "rows": [[cell_to_text(cell) for cell in row] for row in rows]}


def app_now() -> dt.datetime:
    timezone_name = os.environ.get("APP_TIMEZONE", "Asia/Shanghai")
    try:
        return dt.datetime.now(ZoneInfo(timezone_name)).replace(tzinfo=None)
    except ZoneInfoNotFoundError:
        return dt.datetime.utcnow() + dt.timedelta(hours=8)


def now_text() -> str:
    return app_now().strftime("%Y-%m-%d %H:%M:%S")


def parse_datetime(value: str) -> dt.datetime | None:
    value = (value or "").strip()
    if not value:
        return None
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M", "%Y-%m-%d"):
        try:
            parsed = dt.datetime.strptime(value, fmt)
            if fmt == "%Y-%m-%d":
                return parsed.replace(hour=0, minute=0, second=0)
            return parsed
        except ValueError:
            continue
    return None


def row_to_job(row: sqlite3.Row) -> dict[str, object]:
    return {
        "id": row["id"],
        "name": row["name"],
        "enabled": bool(row["enabled"]),
        "steps": json.loads(row["steps_json"] or "[]"),
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def row_to_schedule(row: sqlite3.Row) -> dict[str, object]:
    return {
        "id": row["id"],
        "name": row["name"],
        "jobId": row["job_id"],
        "enabled": bool(row["enabled"]),
        "rule": json.loads(row["rule_json"] or "{}"),
        "startAt": row["start_at"],
        "endAt": row["end_at"],
        "nextRunAt": row["next_run_at"],
        "lastRunAt": row["last_run_at"],
        "lastStatus": row["last_status"],
        "logRetentionDays": row["log_retention_days"],
        "emailOnFail": bool(row["email_on_fail"]),
        "running": bool(row["running"]),
        "createdAt": row["created_at"],
        "updatedAt": row["updated_at"],
    }


def normalize_steps(raw_steps: object) -> list[dict[str, object]]:
    if not isinstance(raw_steps, list):
        return []
    steps: list[dict[str, object]] = []
    for index, raw in enumerate(raw_steps):
        if not isinstance(raw, dict):
            continue
        step_type = str(raw.get("type") or "").strip().lower()
        if step_type not in {"import", "export", "query", "job", "sync"}:
            continue
        steps.append(
            {
                "id": str(raw.get("id") or uuid.uuid4().hex),
                "name": str(raw.get("name") or f"步骤 {index + 1}").strip() or f"步骤 {index + 1}",
                "type": step_type,
                "enabled": raw.get("enabled", True) not in (False, "false", "0", 0, "off"),
                "continueOnError": raw.get("continueOnError", False) in (True, "true", "1", 1, "on"),
                "config": raw.get("config") if isinstance(raw.get("config"), dict) else {},
            }
        )
    return steps


def save_job(payload: dict[str, object]) -> dict[str, object]:
    name = str(payload.get("name") or "").strip()
    if not name:
        raise ValueError("请填写作业名称。")
    steps = normalize_steps(payload.get("steps"))
    if not steps:
        raise ValueError("请至少添加一个子任务。")
    if any(step["type"] == "sync" for step in steps):
        raise ValueError("同步模块尚未开放，暂不能保存同步子任务。")
    job_id = str(payload.get("id") or uuid.uuid4().hex)
    now = now_text()
    with connect_db() as conn:
        old = conn.execute("select created_at from _jobs where id = ?", (job_id,)).fetchone()
        conn.execute(
            """
            insert into _jobs (id, name, enabled, steps_json, created_at, updated_at)
            values (?, ?, ?, ?, ?, ?)
            on conflict(id) do update set
                name = excluded.name,
                enabled = excluded.enabled,
                steps_json = excluded.steps_json,
                updated_at = excluded.updated_at
            """,
            (
                job_id,
                name,
                1 if payload.get("enabled", True) not in (False, "false", "0", 0, "off") else 0,
                json.dumps(steps, ensure_ascii=False),
                old["created_at"] if old else now,
                now,
            ),
        )
        row = conn.execute("select * from _jobs where id = ?", (job_id,)).fetchone()
    return row_to_job(row)


def save_schedule(payload: dict[str, object]) -> dict[str, object]:
    name = str(payload.get("name") or "").strip()
    job_id = str(payload.get("jobId") or payload.get("job_id") or "").strip()
    if not name:
        raise ValueError("请填写任务名称。")
    if not job_id:
        raise ValueError("请选择作业。")
    with connect_db() as conn:
        if not conn.execute("select 1 from _jobs where id = ?", (job_id,)).fetchone():
            raise ValueError("选择的作业不存在。")
    rule = payload.get("rule") if isinstance(payload.get("rule"), dict) else {}
    schedule_id = str(payload.get("id") or uuid.uuid4().hex)
    start_at = str(payload.get("startAt") or payload.get("start_at") or "").replace("T", " ")
    end_at = str(payload.get("endAt") or payload.get("end_at") or "").replace("T", " ")
    retention = max(int(payload.get("logRetentionDays") or payload.get("log_retention_days") or 3), 1)
    enabled = payload.get("enabled", False) in (True, "true", "1", 1, "on")
    next_run = compute_next_run(rule, start_at, end_at, None) if enabled else ""
    now = now_text()
    with connect_db() as conn:
        old = conn.execute("select created_at, last_run_at, last_status from _schedules where id = ?", (schedule_id,)).fetchone()
        conn.execute(
            """
            insert into _schedules (
                id, name, job_id, enabled, rule_json, start_at, end_at, next_run_at,
                last_run_at, last_status, log_retention_days, email_on_fail, running, created_at, updated_at
            )
            values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 0, ?, ?)
            on conflict(id) do update set
                name = excluded.name,
                job_id = excluded.job_id,
                enabled = excluded.enabled,
                rule_json = excluded.rule_json,
                start_at = excluded.start_at,
                end_at = excluded.end_at,
                next_run_at = excluded.next_run_at,
                log_retention_days = excluded.log_retention_days,
                email_on_fail = excluded.email_on_fail,
                updated_at = excluded.updated_at
            """,
            (
                schedule_id,
                name,
                job_id,
                1 if enabled else 0,
                json.dumps(rule, ensure_ascii=False),
                start_at,
                end_at,
                next_run,
                old["last_run_at"] if old else "",
                old["last_status"] if old else "",
                retention,
                1 if payload.get("emailOnFail") in (True, "true", "1", 1, "on") else 0,
                old["created_at"] if old else now,
                now,
            ),
        )
        row = conn.execute("select * from _schedules where id = ?", (schedule_id,)).fetchone()
    return row_to_schedule(row)


def compute_next_run(rule: dict[str, object], start_at: str = "", end_at: str = "", last_run_at: str | None = None, from_time: dt.datetime | None = None) -> str:
    base = from_time or app_now()
    start = parse_datetime(start_at)
    end = parse_datetime(end_at)
    if start and base < start:
        base = start
    if end and base > end:
        return ""
    mode = str(rule.get("mode") or "once")
    last = parse_datetime(last_run_at or "")
    if mode == "once":
        candidate = start or base
        if last:
            return ""
    elif mode == "interval":
        amount = max(int(rule.get("amount") or 1), 1)
        unit = str(rule.get("unit") or "minutes")
        seconds = {"seconds": 1, "minutes": 60, "hours": 3600, "days": 86400}.get(unit, 60) * amount
        if last:
            candidate = last + dt.timedelta(seconds=seconds)
        elif start:
            candidate = start
        else:
            candidate = base + dt.timedelta(seconds=seconds)
        if candidate < base:
            steps = int((base - candidate).total_seconds() // seconds) + 1
            candidate += dt.timedelta(seconds=seconds * steps)
    else:
        time_text = str(rule.get("time") or "09:00:00")
        parts = [int(part) for part in re.findall(r"\d+", time_text)[:3]]
        while len(parts) < 3:
            parts.append(0)
        hour, minute, second = parts[:3]
        if mode == "daily":
            candidate = base.replace(hour=hour, minute=minute, second=second, microsecond=0)
            if candidate <= base:
                candidate += dt.timedelta(days=1)
        elif mode == "weekly":
            weekday = int(rule.get("weekday") or 1)
            weekday = max(1, min(7, weekday)) - 1
            candidate = base.replace(hour=hour, minute=minute, second=second, microsecond=0)
            days = (weekday - candidate.weekday()) % 7
            candidate += dt.timedelta(days=days)
            if candidate <= base:
                candidate += dt.timedelta(days=7)
        elif mode == "monthly":
            day = max(1, min(31, int(rule.get("day") or 1)))
            year, month = base.year, base.month
            while True:
                max_day = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1]
                candidate = dt.datetime(year, month, min(day, max_day), hour, minute, second)
                if candidate > base:
                    break
                month += 1
                if month > 12:
                    year += 1
                    month = 1
        elif mode == "yearly":
            month = max(1, min(12, int(rule.get("month") or 1)))
            day = max(1, min(31, int(rule.get("day") or 1)))
            year = base.year
            while True:
                max_day = [31, 29 if year % 4 == 0 and (year % 100 != 0 or year % 400 == 0) else 28, 31, 30, 31, 30, 31, 31, 30, 31, 30, 31][month - 1]
                candidate = dt.datetime(year, month, min(day, max_day), hour, minute, second)
                if candidate > base:
                    break
                year += 1
        else:
            candidate = base
    if end and candidate > end:
        return ""
    return candidate.strftime("%Y-%m-%d %H:%M:%S")


def collect_local_files(path_text: str) -> list[UploadedFile]:
    path_text = path_text.strip().strip('"')
    if not path_text:
        raise ValueError("请填写导入文件或目录路径。")
    path = Path(path_text)
    if not path.exists():
        raise ValueError(f"路径不存在：{path_text}")
    if path.is_file():
        return [UploadedFile(path.name, path)]
    files = [UploadedFile(item.name, item) for item in sorted(path.iterdir()) if item.is_file() and item.suffix.lower() in SUPPORTED_EXTENSIONS]
    if not files:
        raise ValueError("目录中没有可导入的文件。")
    return files


def execute_query_step(config: dict[str, object]) -> dict[str, object]:
    fields = {key: str(value) for key, value in config.items() if not isinstance(value, (list, dict))}
    sql = str(config.get("sql") or "").strip()
    if not sql:
        raise ValueError("查询 SQL 不能为空。")
    conn = connect_target_db(fields)
    try:
        count = 0
        for statement in [item.strip() for item in sql.split(";") if item.strip()]:
            if target_db_type(fields) == "mysql":
                with conn.cursor() as cursor:
                    cursor.execute(statement)
                    count += cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else len(cursor.fetchall() if cursor.description else [])
            else:
                cursor = conn.execute(statement)
                count += cursor.rowcount if cursor.rowcount and cursor.rowcount > 0 else len(cursor.fetchall() if cursor.description else [])
        if target_db_type(fields) != "mysql":
            conn.commit()
        return {"rows": count}
    finally:
        conn.close()


def execute_import_step(config: dict[str, object]) -> dict[str, object]:
    fields = {key: str(value) for key, value in config.items() if not isinstance(value, (list, dict))}
    files = collect_local_files(str(config.get("path") or config.get("sourcePath") or ""))
    results = [import_uploaded_file(uploaded, fields) for uploaded in files]
    return {
        "files": len(results),
        "rowsWritten": sum(int(item["rowsWritten"]) for item in results),
        "rowsUpdated": sum(int(item["rowsUpdated"]) for item in results),
    }


def run_saved_job(job_id: str, schedule_id: str = "", visited: set[str] | None = None) -> dict[str, object]:
    visited = visited or set()
    if job_id in visited:
        raise ValueError("检测到作业循环引用，已停止执行。")
    visited.add(job_id)
    with connect_db() as conn:
        row = conn.execute("select * from _jobs where id = ?", (job_id,)).fetchone()
    if not row:
        raise ValueError("作业不存在。")
    job = row_to_job(row)
    run_id = uuid.uuid4().hex
    started = dt.datetime.now()
    with connect_db() as conn:
        conn.execute(
            "insert into _job_runs (id, job_id, schedule_id, job_name, started_at, status) values (?, ?, ?, ?, ?, ?)",
            (run_id, job_id, schedule_id, str(job["name"]), started.strftime("%Y-%m-%d %H:%M:%S"), "运行中"),
        )
    status = "成功"
    message = ""
    for index, step in enumerate(job["steps"], start=1):
        if not step.get("enabled", True):
            continue
        step_id = uuid.uuid4().hex
        step_started = dt.datetime.now()
        step_status = "成功"
        step_message = ""
        with connect_db() as conn:
            conn.execute(
                "insert into _job_run_steps (id, run_id, step_index, step_name, step_type, started_at, status) values (?, ?, ?, ?, ?, ?, ?)",
                (step_id, run_id, index, str(step.get("name") or ""), str(step.get("type") or ""), step_started.strftime("%Y-%m-%d %H:%M:%S"), "运行中"),
            )
        try:
            step_type = str(step.get("type") or "")
            config = step.get("config") if isinstance(step.get("config"), dict) else {}
            if step_type == "import":
                result = execute_import_step(config)
                step_message = f"导入 {result['files']} 个文件，写入 {result['rowsWritten']} 行，更新 {result['rowsUpdated']} 行。"
            elif step_type == "export":
                result = run_export_job(config)
                step_message = f"导出 {len(result['files'])} 个文件，{result['rows']} 行。"
            elif step_type == "query":
                result = execute_query_step(config)
                step_message = f"SQL 执行完成，影响/读取 {result['rows']} 行。"
            elif step_type == "job":
                nested = str(config.get("jobId") or "")
                run_saved_job(nested, schedule_id, visited.copy())
                step_message = "子作业执行完成。"
            elif step_type == "sync":
                raise ValueError("同步模块尚未开放。")
            else:
                raise ValueError("不支持的子任务类型。")
        except Exception as exc:
            step_status = "失败"
            step_message = str(exc)
            if not step.get("continueOnError", False):
                status = "失败"
                message = step_message
        finally:
            ended = dt.datetime.now()
            with connect_db() as conn:
                conn.execute(
                    "update _job_run_steps set ended_at = ?, elapsed_ms = ?, status = ?, message = ? where id = ?",
                    (ended.strftime("%Y-%m-%d %H:%M:%S"), int((ended - step_started).total_seconds() * 1000), step_status, step_message, step_id),
                )
        if status == "失败":
            break
    ended = dt.datetime.now()
    if status == "成功":
        message = "作业执行完成。"
    with connect_db() as conn:
        conn.execute(
            "update _job_runs set ended_at = ?, elapsed_ms = ?, status = ?, message = ? where id = ?",
            (ended.strftime("%Y-%m-%d %H:%M:%S"), int((ended - started).total_seconds() * 1000), status, message, run_id),
        )
    return {"id": run_id, "jobId": job_id, "status": status, "message": message}


def prune_job_logs(retention_days: int) -> None:
    cutoff = (dt.datetime.now() - dt.timedelta(days=max(retention_days, 1))).strftime("%Y-%m-%d %H:%M:%S")
    with connect_db() as conn:
        old_ids = [row["id"] for row in conn.execute("select id from _job_runs where started_at < ?", (cutoff,)).fetchall()]
        if old_ids:
            placeholders = ",".join("?" for _ in old_ids)
            conn.execute(f"delete from _job_run_steps where run_id in ({placeholders})", old_ids)
            conn.execute(f"delete from _job_runs where id in ({placeholders})", old_ids)


def dispatch_due_schedules() -> None:
    now = now_text()
    with connect_db() as conn:
        rows = conn.execute(
            "select * from _schedules where enabled = 1 and running = 0 and next_run_at != '' and next_run_at <= ?",
            (now,),
        ).fetchall()
        for row in rows:
            conn.execute("update _schedules set running = 1 where id = ?", (row["id"],))
    for row in rows:
        threading.Thread(target=run_schedule_once, args=(row["id"],), daemon=True).start()


def run_schedule_once(schedule_id: str) -> None:
    with connect_db() as conn:
        row = conn.execute("select * from _schedules where id = ?", (schedule_id,)).fetchone()
    if not row:
        return
    schedule = row_to_schedule(row)
    status = "成功"
    try:
        prune_job_logs(int(schedule["logRetentionDays"]))
        result = run_saved_job(str(schedule["jobId"]), schedule_id)
        status = str(result["status"])
    except Exception as exc:
        status = "失败"
        with connect_db() as conn:
            conn.execute(
                "insert into _job_runs (id, job_id, schedule_id, job_name, started_at, ended_at, status, message) values (?, ?, ?, ?, ?, ?, ?, ?)",
                (uuid.uuid4().hex, str(schedule["jobId"]), schedule_id, str(schedule["name"]), now_text(), now_text(), "失败", str(exc)),
            )
    rule = schedule["rule"] if isinstance(schedule["rule"], dict) else {}
    next_run = compute_next_run(rule, str(schedule["startAt"]), str(schedule["endAt"]), now_text())
    enabled = 1 if next_run else 0
    with connect_db() as conn:
        conn.execute(
            "update _schedules set running = 0, enabled = ?, last_run_at = ?, last_status = ?, next_run_at = ?, updated_at = ? where id = ?",
            (enabled, now_text(), status, next_run, now_text(), schedule_id),
        )


def scheduler_loop(stop_event: threading.Event) -> None:
    while not stop_event.is_set():
        try:
            dispatch_due_schedules()
        except Exception:
            pass
        stop_event.wait(5)


def log_import(
    conn: sqlite3.Connection,
    file_name: str,
    table_name: str,
    mode: str,
    rows_read: int,
    rows_written: int,
    rows_updated: int,
    rows_skipped: int,
    status: str,
    message: str,
) -> None:
    if parse_bool({"disableLog": "false"}, "disableLog", False):
        return
    conn.execute(
        """
        insert into _import_logs
        (id, created_at, file_name, table_name, mode, rows_read, rows_written, rows_updated, rows_skipped, status, message)
        values (?, datetime('now', 'localtime'), ?, ?, ?, ?, ?, ?, ?, ?, ?)
        """,
        (str(uuid.uuid4()), file_name, table_name, mode, rows_read, rows_written, rows_updated, rows_skipped, status, message),
    )


def import_uploaded_file(uploaded: UploadedFile, fields: dict[str, str]) -> dict[str, object]:
    if fields.get("sheetMode") == "all" and uploaded.path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
        results = []
        original_table_name = fields.get("tableName", "")
        for tabular in read_tabular_tasks(uploaded.path, fields):
            per_sheet = dict(fields)
            per_sheet["sheetMode"] = "specified"
            per_sheet["sheetFilterMode"] = "name"
            per_sheet["sheetName"] = tabular.selected_sheet
            if original_table_name:
                per_sheet["tableName"] = f"{original_table_name}_{tabular.selected_sheet}"
            else:
                per_sheet["tableNameRule"] = "sheet"
            results.append(import_uploaded_file(uploaded, per_sheet))
        return {
            "fileName": uploaded.filename,
            "tableName": results[0]["tableName"] if results else "",
            "columns": results[0]["columns"] if results else [],
            "rowsRead": sum(int(item["rowsRead"]) for item in results),
            "rowsWritten": sum(int(item["rowsWritten"]) for item in results),
            "rowsUpdated": sum(int(item["rowsUpdated"]) for item in results),
            "rowsSkipped": sum(int(item["rowsSkipped"]) for item in results),
            "message": f"已导入 {len(results)} 个 Sheet",
            "sheetResults": results,
        }

    tabular = read_tabular_file(uploaded.path, fields)
    columns, rows, match_keys, skipped = build_target_data(tabular, fields, uploaded.filename)
    table_name = normalize_target_name(uploaded, tabular, fields)
    mode = fields.get("importMode", "append")
    if mode not in {"append", "update", "overwrite", "rebuild"}:
        raise ValueError("导入模式只能是追加、更新、覆盖或重建。")

    conn = connect_target_db(fields)
    try:
        duplicate_mode = fields.get("duplicateTableMode", "same")
        if target_table_exists(conn, table_name, fields) and mode == "append" and duplicate_mode == "skip":
            if not parse_bool(fields, "disableLog", False):
                with connect_db() as log_conn:
                    log_import(log_conn, uploaded.filename, table_name, mode, len(tabular.rows), 0, 0, len(rows), "成功", "目标表重复，已按配置跳过")
            return {
                "fileName": uploaded.filename,
                "tableName": table_name,
                "columns": columns,
                "rowsRead": len(tabular.rows),
                "rowsWritten": 0,
                "rowsUpdated": 0,
                "rowsSkipped": len(rows),
                "message": "目标表重复，已跳过",
            }
        if target_table_exists(conn, table_name, fields) and mode == "append" and duplicate_mode == "suffix":
            base = table_name
            index = 2
            while target_table_exists(conn, table_name, fields):
                table_name = sanitize_identifier(f"{base}_{index}", "import_table")
                index += 1

        cp_key = checkpoint_key(uploaded, table_name, fields)
        resume_offset = get_checkpoint(cp_key) if parse_bool(fields, "resumeImport", False) and mode in {"append", "rebuild", "overwrite"} else 0
        target_create_or_expand_table(conn, table_name, columns, rows, mode == "rebuild" and resume_offset == 0, parse_bool(fields, "autoExpand", True), fields)
        existing = target_existing_columns(conn, table_name, fields)
        if any(column not in existing for column in columns):
            raise ValueError("目标表字段与当前导入字段不一致。")

        if mode == "overwrite" and resume_offset == 0:
            sql = f"delete from {db_quote(table_name, fields)}"
            if target_db_type(fields) == "mysql":
                with conn.cursor() as cursor:
                    cursor.execute(sql)
            else:
                conn.execute(sql)

        active_rows = rows[resume_offset:] if resume_offset else rows

        def progress(done: int) -> None:
            if parse_bool(fields, "resumeImport", False):
                if target_db_type(fields) == "sqlite":
                    conn.commit()
                set_checkpoint(cp_key, resume_offset + done)

        if mode == "update":
            written, updated = target_update_rows(conn, table_name, columns, rows, match_keys, fields)
        else:
            if fields.get("writeMode") == "load" and target_db_type(fields) == "mysql":
                try:
                    written = mysql_load_rows(conn, table_name, columns, active_rows, fields, progress)
                except Exception:
                    written = target_insert_rows(conn, table_name, columns, active_rows, fields, progress)
            elif fields.get("writeMode") == "parallel":
                conn.commit()
                written = target_insert_rows_parallel(table_name, columns, active_rows, fields, progress)
            else:
                written = target_insert_rows(conn, table_name, columns, active_rows, fields, progress)
            updated = 0

        target_execute_sql_batch(conn, fields.get("customSql", ""), "自定义 SQL", fields)
        target_execute_sql_batch(conn, fields.get("afterEachSql", ""), "每次导入成功后 SQL", fields)
        if target_db_type(fields) != "mysql" or fields.get("commitMode") != "auto":
            conn.commit()
        if not parse_bool(fields, "disableLog", False):
            with connect_db() as log_conn:
                log_import(log_conn, uploaded.filename, table_name, mode, len(tabular.rows), written, updated, skipped, "成功", "导入完成")
        if parse_bool(fields, "resumeImport", False):
            clear_checkpoint(cp_key)
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()

    if parse_bool(fields, "deleteAfterSuccess", False):
        uploaded.path.unlink(missing_ok=True)

    return {
        "fileName": uploaded.filename,
        "tableName": table_name,
        "columns": columns,
        "rowsRead": len(tabular.rows),
        "rowsWritten": written,
        "rowsUpdated": updated,
        "rowsSkipped": skipped,
        "message": "导入完成",
    }

def parse_multipart(handler: SimpleHTTPRequestHandler) -> tuple[dict[str, str], list[UploadedFile]]:
    content_type = handler.headers.get("Content-Type", "")
    if "multipart/form-data" not in content_type:
        raise ValueError("请求格式错误，需要 multipart/form-data。")
    boundary_match = re.search(r"boundary=(.+)", content_type)
    if not boundary_match:
        raise ValueError("上传请求缺少 boundary。")
    boundary = boundary_match.group(1).strip('"')
    length = int(handler.headers.get("Content-Length", "0"))
    body = handler.rfile.read(length)
    parts = body.split(("--" + boundary).encode("utf-8"))
    fields: dict[str, str] = {}
    uploaded_files: list[UploadedFile] = []

    for part in parts:
        part = part.strip(b"\r\n")
        if not part or part == b"--":
            continue
        header_blob, _, content = part.partition(b"\r\n\r\n")
        headers = header_blob.decode("utf-8", errors="replace")
        name_match = re.search(r'name="([^"]+)"', headers)
        if not name_match:
            continue
        name = name_match.group(1)
        filename_match = re.search(r'filename="([^"]*)"', headers)
        content = content.rstrip(b"\r\n")
        if filename_match and filename_match.group(1):
            original = Path(filename_match.group(1)).name
            target = UPLOADS / f"{int(time.time())}_{uuid.uuid4().hex}_{original}"
            target.write_bytes(content)
            uploaded_files.append(UploadedFile(filename=original, path=target))
        else:
            fields[name] = content.decode("utf-8", errors="replace")
    return fields, uploaded_files


def read_json_body(handler: SimpleHTTPRequestHandler) -> dict[str, object]:
    length = int(handler.headers.get("Content-Length", "0") or "0")
    if length <= 0:
        return {}
    raw = handler.rfile.read(length).decode("utf-8")
    payload = json.loads(raw) if raw.strip() else {}
    if not isinstance(payload, dict):
        raise ValueError("请求内容格式不正确。")
    return payload


def public_auth_enabled() -> bool:
    enabled_value = os.environ.get("APP_AUTH_ENABLED", "").strip().lower()
    return enabled_value in {"1", "true", "yes", "on"} and bool(os.environ.get("ADMIN_PASSWORD", "").strip())


def check_basic_auth(header_value: str) -> bool:
    admin_user = os.environ.get("ADMIN_USER", "admin")
    admin_password = os.environ.get("ADMIN_PASSWORD", "")
    if not admin_password:
        return True
    if not header_value.startswith("Basic "):
        return False
    try:
        decoded = base64.b64decode(header_value[6:].strip()).decode("utf-8")
    except Exception:
        return False
    user, separator, password = decoded.partition(":")
    if not separator:
        return False
    return hmac.compare_digest(user, admin_user) and hmac.compare_digest(password, admin_password)


def bind_host() -> str:
    if os.environ.get("RAILWAY_ENVIRONMENT") or os.environ.get("RAILWAY_ENVIRONMENT_NAME"):
        return "0.0.0.0"
    return os.environ.get("HOST", "127.0.0.1")


class ImportPrototypeHandler(SimpleHTTPRequestHandler):
    def require_auth(self) -> bool:
        if not public_auth_enabled():
            return True
        if urlparse(self.path).path == "/api/ping":
            return True
        if check_basic_auth(self.headers.get("Authorization", "")):
            return True
        self.send_response(HTTPStatus.UNAUTHORIZED)
        self.send_header("WWW-Authenticate", 'Basic realm="Data Converter"')
        self.send_header("Content-Type", "text/plain; charset=utf-8")
        self.end_headers()
        self.wfile.write("Authentication required.".encode("utf-8"))
        return False

    def end_headers(self) -> None:
        self.send_header("Cache-Control", "no-store, no-cache, must-revalidate, max-age=0")
        self.send_header("Pragma", "no-cache")
        self.send_header("Expires", "0")
        super().end_headers()

    def translate_path(self, path: str) -> str:
        parsed = urlparse(path)
        request_path = parsed.path
        if request_path == "/":
            request_path = "/index.html"
        base = PUBLIC.resolve()
        target = (base / request_path.lstrip("/")).resolve()
        try:
            target.relative_to(base)
        except ValueError:
            return str(base / "__missing__")
        return str(target)

    def do_GET(self) -> None:
        if not self.require_auth():
            return
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/tables":
                self.handle_tables()
                return
            if parsed.path == "/api/target-tables":
                self.handle_target_tables(parsed.query)
                return
            if parsed.path == "/api/table":
                self.handle_table_preview(parsed.query)
                return
            if parsed.path == "/api/logs":
                self.handle_logs()
                return
            if parsed.path == "/api/ping":
                json_response(self, {"ok": True, "version": "export-v2-download"})
                return
            if parsed.path == "/api/storage/status":
                self.handle_storage_status()
                return
            if parsed.path == "/api/connections":
                self.handle_connections()
                return
            if parsed.path == "/api/jobs":
                self.handle_jobs()
                return
            if parsed.path == "/api/schedules":
                self.handle_schedules()
                return
            if parsed.path == "/api/job-runs":
                self.handle_job_runs(parsed.query)
                return
            if parsed.path == "/api/export/sources":
                self.handle_export_sources(parsed.query)
                return
            if parsed.path == "/api/export/download":
                self.handle_export_download(parsed.query)
                return
        except Exception as exc:
            error_response(self, str(exc), HTTPStatus.BAD_REQUEST)
            return
        return super().do_GET()

    def do_POST(self) -> None:
        if not self.require_auth():
            return
        try:
            if self.path == "/api/preview":
                self.handle_preview()
                return
            if self.path == "/api/import":
                self.handle_import()
                return
            if self.path == "/api/connections/test":
                self.handle_connection_test()
                return
            if self.path == "/api/connections":
                self.handle_connection_save()
                return
            if self.path == "/api/export/preview":
                self.handle_export_preview()
                return
            if self.path == "/api/export/run":
                self.handle_export_run()
                return
            if self.path == "/api/jobs":
                self.handle_job_save()
                return
            if self.path == "/api/jobs/run":
                self.handle_job_run()
                return
            if self.path == "/api/schedules":
                self.handle_schedule_save()
                return
            if self.path == "/api/schedules/start":
                self.handle_schedule_state(True)
                return
            if self.path == "/api/schedules/pause":
                self.handle_schedule_state(False)
                return
            error_response(self, "未知接口。", HTTPStatus.NOT_FOUND)
        except Exception as exc:
            error_response(self, str(exc), HTTPStatus.BAD_REQUEST)

    def do_DELETE(self) -> None:
        if not self.require_auth():
            return
        parsed = urlparse(self.path)
        try:
            if parsed.path == "/api/connections":
                self.handle_connection_delete(parsed.query)
                return
            if parsed.path == "/api/jobs":
                self.handle_job_delete(parsed.query)
                return
            if parsed.path == "/api/schedules":
                self.handle_schedule_delete(parsed.query)
                return
            error_response(self, "未知接口。", HTTPStatus.NOT_FOUND)
        except Exception as exc:
            error_response(self, str(exc), HTTPStatus.BAD_REQUEST)

    def guess_type(self, path: str) -> str:
        if path.endswith(".js"):
            return "text/javascript; charset=utf-8"
        if path.endswith(".css"):
            return "text/css; charset=utf-8"
        if path.endswith(".html"):
            return "text/html; charset=utf-8"
        return mimetypes.guess_type(path)[0] or "application/octet-stream"

    def handle_preview(self) -> None:
        fields, uploaded_files = parse_multipart(self)
        if not uploaded_files:
            raise ValueError("请选择要预览的文件。")
        uploaded = uploaded_files[0]
        tabular = read_tabular_file(uploaded.path, fields)
        json_response(
            self,
            {
                "ok": True,
                "fileName": uploaded.filename,
                "suggestedTable": sanitize_identifier(Path(uploaded.filename).stem, "import_table"),
                "columns": tabular.columns,
                "preview": tabular.rows[:MAX_PREVIEW_ROWS],
                "totalRows": len(tabular.rows),
                "sheets": tabular.sheets,
                "selectedSheet": tabular.selected_sheet,
            },
        )

    def handle_import(self) -> None:
        fields, uploaded_files = parse_multipart(self)
        if not uploaded_files:
            raise ValueError("请选择要导入的文件。")

        results = []
        failures = []
        export_path = ""
        with connect_db() as log_conn:
            if parse_bool(fields, "clearLogBeforeImport", False):
                log_conn.execute("delete from _import_logs")

        target_conn = connect_target_db(fields)
        try:
            target_execute_sql_batch(target_conn, fields.get("beforeAllSql", ""), "全部导入开始前 SQL", fields)
            if target_db_type(fields) != "mysql" or fields.get("commitMode") != "auto":
                target_conn.commit()
        finally:
            target_conn.close()

        for uploaded in uploaded_files:
            try:
                results.append(import_uploaded_file(uploaded, fields))
            except Exception as exc:
                failures.append({"fileName": uploaded.filename, "error": str(exc)})

        target_conn = connect_target_db(fields)
        try:
            target_execute_sql_batch(target_conn, fields.get("afterAllSql", ""), "全部导入结束后 SQL", fields)
            export_path = target_export_query_to_excel(target_conn, fields.get("afterQuerySql", ""), fields.get("afterQueryExport", ""), fields)
            if target_db_type(fields) != "mysql" or fields.get("commitMode") != "auto":
                target_conn.commit()
        finally:
            target_conn.close()

        if failures and not results:
            raise ValueError(failures[0]["error"])

        json_response(
            self,
            {
                "ok": True,
                "tableName": results[0]["tableName"] if results else "",
                "columns": results[0]["columns"] if results else [],
                "summary": {
                    "totalFiles": len(uploaded_files),
                    "successFiles": len(results),
                    "failedFiles": len(failures),
                    "rowsRead": sum(int(item["rowsRead"]) for item in results),
                    "rowsWritten": sum(int(item["rowsWritten"]) for item in results),
                    "rowsUpdated": sum(int(item["rowsUpdated"]) for item in results),
                    "rowsSkipped": sum(int(item["rowsSkipped"]) for item in results),
                },
                "results": results,
                "failures": failures,
                "exportPath": export_path,
                "message": f"成功导入 {len(results)} 个文件，失败 {len(failures)} 个文件。",
            },
        )

    def handle_tables(self) -> None:
        with connect_db() as conn:
            rows = conn.execute(
                """
                select name from sqlite_master
                where type = 'table' and name not like 'sqlite_%' and name not like '\\_%' escape '\\'
                order by name
                """
            ).fetchall()
        json_response(self, {"ok": True, "tables": [row["name"] for row in rows]})

    def handle_target_tables(self, query: str) -> None:
        raw = {key: values[-1] for key, values in parse_qs(query).items()}
        fields = {key: str(value) for key, value in raw.items()}
        sources = export_sources(fields)
        json_response(self, {"ok": True, "tables": [str(item["name"]) for item in sources if item.get("type") != "VIEW"]})

    def handle_table_preview(self, query: str) -> None:
        params = parse_qs(query)
        table_name = sanitize_identifier(params.get("name", [""])[0], "")
        if not table_name:
            raise ValueError("缺少表名。")
        with connect_db() as conn:
            columns = existing_columns(conn, table_name)
            if not columns:
                raise ValueError("表不存在或没有字段。")
            rows = conn.execute(f"select * from {quote_identifier(table_name)} limit 50").fetchall()
            count = conn.execute(f"select count(*) as total from {quote_identifier(table_name)}").fetchone()["total"]
        json_response(
            self,
            {
                "ok": True,
                "tableName": table_name,
                "columns": columns,
                "rows": [[cell_to_text(row[column]) for column in columns] for row in rows],
                "totalRows": count,
            },
        )

    def handle_logs(self) -> None:
        with connect_db() as conn:
            rows = conn.execute(
                """
                select created_at, file_name, table_name, mode, rows_read, rows_written, rows_updated, rows_skipped, status, message
                from _import_logs
                order by created_at desc
                limit 50
                """
            ).fetchall()
        json_response(self, {"ok": True, "logs": [dict(row) for row in rows]})

    def handle_storage_status(self) -> None:
        ensure_dirs()
        volume_mount = os.environ.get("RAILWAY_VOLUME_MOUNT_PATH", "").strip()
        write_test_path = DATA / ".storage-write-test"
        write_ok = False
        write_error = ""
        try:
            write_test_path.write_text(now_text(), encoding="utf-8")
            write_ok = write_test_path.exists()
        except Exception as exc:
            write_error = str(exc)
        def inside_volume(path: Path) -> bool:
            if not volume_mount:
                return False
            try:
                path.resolve().relative_to(Path(volume_mount).resolve())
                return True
            except ValueError:
                return False
        json_response(
            self,
            {
                "ok": True,
                "railwayVolumeMountPath": volume_mount,
                "persistentReady": bool(volume_mount) and inside_volume(DB_PATH) and write_ok,
                "dataDir": str(DATA),
                "uploadsDir": str(UPLOADS),
                "exportsDir": str(EXPORTS),
                "databasePath": str(DB_PATH),
                "databaseExists": DB_PATH.exists(),
                "writeTestOk": write_ok,
                "writeTestError": write_error,
            },
        )

    def handle_connections(self) -> None:
        with connect_db() as conn:
            rows = conn.execute("select * from _db_connections order by updated_at desc, name").fetchall()
        json_response(self, {"ok": True, "connections": [connection_public(row) for row in rows]})

    def handle_connection_test(self) -> None:
        payload = read_json_body(self)
        result = test_mysql_connection(payload)
        json_response(
            self,
            {
                "ok": True,
                "message": "连接成功。",
                "version": result["version"],
                "databases": result["databases"],
            },
        )

    def handle_connection_save(self) -> None:
        payload = read_json_body(self)
        record = normalize_connection_payload(payload)
        now = dt.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        with connect_db() as conn:
            old = conn.execute("select created_at, password from _db_connections where id = ?", (record["id"],)).fetchone()
            raw_password = str(payload.get("password") or "")
            stored_password = old["password"] if old and not raw_password else encode_secret(str(record["password"]))
            conn.execute(
                """
                insert into _db_connections (
                    id, name, db_type, host, port, user_name, password, db_name, charset,
                    ssl_enabled, ssl_ca, ssl_cert, ssl_key, created_at, updated_at
                )
                values (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                on conflict(id) do update set
                    name = excluded.name,
                    db_type = excluded.db_type,
                    host = excluded.host,
                    port = excluded.port,
                    user_name = excluded.user_name,
                    password = excluded.password,
                    db_name = excluded.db_name,
                    charset = excluded.charset,
                    ssl_enabled = excluded.ssl_enabled,
                    ssl_ca = excluded.ssl_ca,
                    ssl_cert = excluded.ssl_cert,
                    ssl_key = excluded.ssl_key,
                    updated_at = excluded.updated_at
                """,
                (
                    record["id"],
                    record["name"],
                    record["db_type"],
                    record["host"],
                    record["port"],
                    record["user_name"],
                    stored_password,
                    record["db_name"],
                    record["charset"],
                    record["ssl_enabled"],
                    record["ssl_ca"],
                    record["ssl_cert"],
                    record["ssl_key"],
                    old["created_at"] if old else now,
                    now,
                ),
            )
            row = conn.execute("select * from _db_connections where id = ?", (record["id"],)).fetchone()
        json_response(self, {"ok": True, "connection": connection_public(row)})

    def handle_connection_delete(self, query: str) -> None:
        params = parse_qs(query)
        connection_id = params.get("id", [""])[0]
        if not connection_id:
            raise ValueError("缺少连接编号。")
        with connect_db() as conn:
            conn.execute("delete from _db_connections where id = ?", (connection_id,))
        json_response(self, {"ok": True})

    def handle_jobs(self) -> None:
        with connect_db() as conn:
            rows = conn.execute("select * from _jobs order by updated_at desc, name").fetchall()
        json_response(self, {"ok": True, "jobs": [row_to_job(row) for row in rows]})

    def handle_job_save(self) -> None:
        job = save_job(read_json_body(self))
        json_response(self, {"ok": True, "job": job})

    def handle_job_run(self) -> None:
        payload = read_json_body(self)
        job_id = str(payload.get("id") or payload.get("jobId") or "").strip()
        schedule_id = str(payload.get("scheduleId") or "").strip()
        if not job_id:
            raise ValueError("缺少作业编号。")
        result = run_saved_job(job_id, schedule_id)
        json_response(self, {"ok": True, "run": result})

    def handle_job_delete(self, query: str) -> None:
        params = parse_qs(query)
        job_id = params.get("id", [""])[0]
        if not job_id:
            raise ValueError("缺少作业编号。")
        with connect_db() as conn:
            conn.execute("delete from _jobs where id = ?", (job_id,))
            conn.execute("delete from _schedules where job_id = ?", (job_id,))
        json_response(self, {"ok": True})

    def handle_schedules(self) -> None:
        with connect_db() as conn:
            rows = conn.execute("select * from _schedules order by updated_at desc, name").fetchall()
        json_response(self, {"ok": True, "schedules": [row_to_schedule(row) for row in rows]})

    def handle_schedule_save(self) -> None:
        schedule = save_schedule(read_json_body(self))
        json_response(self, {"ok": True, "schedule": schedule})

    def handle_schedule_state(self, enabled: bool) -> None:
        payload = read_json_body(self)
        schedule_id = str(payload.get("id") or "").strip()
        if not schedule_id:
            raise ValueError("缺少定时任务编号。")
        with connect_db() as conn:
            row = conn.execute("select * from _schedules where id = ?", (schedule_id,)).fetchone()
            if not row:
                raise ValueError("定时任务不存在。")
            schedule = row_to_schedule(row)
            next_run = compute_next_run(schedule["rule"], str(schedule["startAt"]), str(schedule["endAt"]), None) if enabled else ""
            conn.execute(
                "update _schedules set enabled = ?, running = 0, next_run_at = ?, updated_at = ? where id = ?",
                (1 if enabled else 0, next_run, now_text(), schedule_id),
            )
            row = conn.execute("select * from _schedules where id = ?", (schedule_id,)).fetchone()
        json_response(self, {"ok": True, "schedule": row_to_schedule(row)})

    def handle_schedule_delete(self, query: str) -> None:
        params = parse_qs(query)
        schedule_id = params.get("id", [""])[0]
        if not schedule_id:
            raise ValueError("缺少定时任务编号。")
        with connect_db() as conn:
            conn.execute("delete from _schedules where id = ?", (schedule_id,))
        json_response(self, {"ok": True})

    def handle_job_runs(self, query: str) -> None:
        params = parse_qs(query)
        job_id = params.get("jobId", [""])[0]
        schedule_id = params.get("scheduleId", [""])[0]
        where = []
        args: list[object] = []
        if job_id:
            where.append("job_id = ?")
            args.append(job_id)
        if schedule_id:
            where.append("schedule_id = ?")
            args.append(schedule_id)
        clause = (" where " + " and ".join(where)) if where else ""
        with connect_db() as conn:
            runs = conn.execute(f"select * from _job_runs{clause} order by started_at desc limit 80", args).fetchall()
            run_ids = [row["id"] for row in runs]
            steps_by_run: dict[str, list[dict[str, object]]] = {run_id: [] for run_id in run_ids}
            if run_ids:
                placeholders = ",".join("?" for _ in run_ids)
                steps = conn.execute(f"select * from _job_run_steps where run_id in ({placeholders}) order by step_index", run_ids).fetchall()
                for step in steps:
                    steps_by_run[step["run_id"]].append(dict(step))
        json_response(self, {"ok": True, "runs": [{**dict(row), "steps": steps_by_run.get(row["id"], [])} for row in runs]})

    def handle_export_sources(self, query: str) -> None:
        params = parse_qs(query)
        fields = {
            "connectionId": params.get("connectionId", [""])[0],
            "targetDbType": params.get("targetDbType", ["mysql"])[0],
        }
        json_response(self, {"ok": True, "sources": export_sources(fields)})

    def handle_export_preview(self) -> None:
        payload = read_json_body(self)
        preview = preview_export_job(payload)
        json_response(self, {"ok": True, **preview})

    def handle_export_run(self) -> None:
        payload = read_json_body(self)
        result = run_export_job(payload)
        json_response(
            self,
            {
                "ok": True,
                "files": result["files"],
                "downloadUrls": ["/api/export/download?name=" + quote(Path(path).name) for path in result["files"]],
                "rows": result["rows"],
                "elapsedMs": result["elapsedMs"],
                "message": f"导出完成：{len(result['files'])} 个文件，{result['rows']} 行。",
            },
        )

    def handle_export_download(self, query: str) -> None:
        params = parse_qs(query)
        name = Path(params.get("name", [""])[0]).name
        if not name:
            raise ValueError("缺少文件名。")
        path = EXPORTS / name
        if not path.exists() or not path.is_file():
            raise ValueError("导出文件不存在。")
        ascii_stem = re.sub(r"[^A-Za-z0-9_-]+", "_", path.stem, flags=re.ASCII).strip("_") or "export"
        ascii_suffix = re.sub(r"[^A-Za-z0-9.]+", "", path.suffix, flags=re.ASCII)
        ascii_name = f"{ascii_stem}{ascii_suffix}"
        encoded_name = quote(name)
        self.send_response(HTTPStatus.OK)
        self.send_header("Content-Type", mimetypes.guess_type(str(path))[0] or "application/octet-stream")
        self.send_header("Content-Disposition", f"attachment; filename=\"{ascii_name}\"; filename*=UTF-8''{encoded_name}")
        self.send_header("Content-Length", str(path.stat().st_size))
        self.end_headers()
        with path.open("rb") as file:
            while True:
                chunk = file.read(1024 * 1024)
                if not chunk:
                    break
                self.wfile.write(chunk)


def main() -> None:
    ensure_dirs()
    port = int(os.environ.get("PORT", "8765"))
    host = bind_host()
    server = ThreadingHTTPServer((host, port), ImportPrototypeHandler)
    stop_event = threading.Event()
    scheduler = threading.Thread(target=scheduler_loop, args=(stop_event,), daemon=True)
    scheduler.start()
    print(f"Import prototype running at http://{host}:{port}", flush=True)
    try:
        server.serve_forever()
    except KeyboardInterrupt:
        print("\nStopping import prototype...")
    finally:
        stop_event.set()
        server.server_close()


if __name__ == "__main__":
    main()
